from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import time
import traceback
import getpass
from openpyxl import load_workbook
import sys

# 크롤링할 최대 게시글 수 설정
max_posts = 10 

# 엑셀 파일 경로 및 워크시트 이름 설정
excel_file = r'C:\Users\PHJ\output\개인정보 운영대장.xlsx'
worksheet_name = '개인정보 추출 및 이용 관리'

# 웹드라이버 설정
driver = webdriver.Chrome()

try:
    # 로그인 페이지로 이동
    driver.get('https://gw.com2us.com/')

    # 로그인 처리
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, 'Username')))
    username_input = driver.find_element(By.ID, 'Username')
    password_input = driver.find_element(By.ID, 'Password')

    username = input('아이디를 입력하세요: ')
    password = getpass.getpass('비밀번호를 입력하세요: ')

    username_input.send_keys(username)
    password_input.send_keys(password)

    login_button = driver.find_element(By.CLASS_NAME, 'btnLogin')
    login_button.click()

    # 로그인 성공 여부 확인
    WebDriverWait(driver, 30).until(EC.url_changes('https://gw.com2us.com/'))
    current_url = driver.current_url
    print(f"로그인 후 현재 URL: {current_url}")

    if 'login' in current_url.lower():
        print("로그인에 실패하였습니다.")
        driver.quit()
        sys.exit()

    # 결재 > 부서함 페이지로 이동
    driver.get('https://gw.com2us.com/emate_appro/appro_complete_2024_link.nsf/wfmViaView?readform&viewname=view055&vctype=a')
    print(f"페이지 이동 후 현재 URL: {driver.current_url}")

    # 페이지 로딩 대기: 검색창 요소가 로드될 때까지 대기
    try:
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, 'searchtext')))
    except Exception as e:
        print("검색창을 찾을 수 없습니다.")
        print(e)
        driver.quit()
        sys.exit()

    # 검색창에 '개인정보 추출 신청서' 입력
    search_input = driver.find_element(By.ID, 'searchtext')
    search_input.clear()
    search_input.send_keys('개인정보 추출 신청서')

    # 검색 버튼 클릭
    try:
        # 검색 버튼 이미지의 src 속성이 'btn_search_board.gif'인 img 태그 찾기
        search_button = driver.find_element(By.XPATH, '//img[@class="inbtn" and contains(@src, "btn_search_board.gif")]')
        search_button.click()
    except Exception as e:
        print("검색 버튼을 찾을 수 없습니다.")
        print(e)
        driver.quit()
        sys.exit()

    # 검색 결과 로딩 대기
    time.sleep(5)  # 필요에 따라 조정

    # 게시글 목록 가져오기
    posts = driver.find_elements(By.XPATH, '//tr[contains(@class, "dhx_skyblue")]')
    print(f"게시글 수: {len(posts)}")

    data_list = []

    # 크롤링할 게시글 수 계산
    num_posts_to_crawl = min(len(posts), max_posts)

    for i in range(num_posts_to_crawl):
        # 게시글 목록을 다시 가져옵니다 (페이지가 동적으로 변할 수 있으므로)
        posts = driver.find_elements(By.CSS_SELECTOR, 'tr[class*="dhx_skyblue"]')
        if i >= len(posts):
            print(f"게시글 수가 예상보다 적습니다. 현재 인덱스: {i}, 게시글 수: {len(posts)}")
            break
        post = posts[i]

        try:
            # 해당 행의 모든 td 요소를 가져옵니다.
            tds = post.find_elements(By.TAG_NAME, 'td')

            # 결재일 추출
            결재일_td = tds[5]  # 결재일이 6번째 컬럼에 위치
            결재일_text = 결재일_td.text.strip()

            # 년, 월, 일 추출
            년, 월, 일 = 결재일_text.split('-')
            월 = str(int(월))  # 0을 제거
            일 = str(int(일))  # 0을 제거

            # 신청자 추출
            신청자_td = tds[4]  # 신청자가 5번째 컬럼에 위치
            신청자_span = 신청자_td.find_element(By.TAG_NAME, 'span')
            신청자 = 신청자_span.text.strip()

            print(f"게시글 {i+1}/{num_posts_to_crawl} - 결재일: {결재일_text}, 신청자: {신청자}")

        except Exception as e:
            print(f"목록에서 데이터 추출 중 오류 발생 (게시글 {i+1}): {e}")
            traceback.print_exc()
            continue  # 오류 발생 시 다음 게시글로 이동

        # 요소가 화면에 보이도록 스크롤합니다.
        driver.execute_script("arguments[0].scrollIntoView();", post)

        # 클릭 가능할 때까지 대기합니다.
        try:
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable(post))
        except Exception as e:
            print(f"게시글을 클릭할 수 없습니다 (게시글 {i+1}): {e}")
            traceback.print_exc()
            continue

        # 게시글 클릭하여 상세 페이지 열기
        try:
            post.click()
        except Exception as e:
            print(f"게시글 클릭 중 오류 발생 (게시글 {i+1}): {e}")
            traceback.print_exc()
            continue

        # 새로운 창으로 전환
        try:
            WebDriverWait(driver, 10).until(EC.number_of_windows_to_be(2))
            driver.switch_to.window(driver.window_handles[-1])
        except Exception as e:
            print(f"새 창으로 전환 중 오류 발생 (게시글 {i+1}): {e}")
            traceback.print_exc()
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
            continue

        # 상세 페이지 로딩 대기_서류명이 '개인정보 추출 신청서'인지 확인하기 위함
        try:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'AppLineArea')))
        except Exception as e:
            print(f"상세 페이지 로딩 중 오류 발생 (게시글 {i+1}): {e}")
            traceback.print_exc()
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
            continue

        try:
            # 상세 페이지에서 제목 확인_서류명이 '개인정보 추출 신청서'인지
            h2_element = driver.find_element(By.CSS_SELECTOR, '#AppLineArea h2')
            h2_text = h2_element.text.strip()

            # 제목이 '개인정보 추출 신청서'가 아닌 경우 건너뜀
            if h2_text != '개인정보 추출 신청서':
                print(f"게시글 {i+1} - 제목이 '개인정보 추출 신청서'가 아닙니다: {h2_text}")
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                continue  # 다음 게시글로 이동

            # 현재 창 제목 출력
            print(f"게시글 {i+1} - 현재 창 제목: {driver.title}")

            # 법인명 추출
            법인명_element = driver.find_elements(By.ID, 'titleLabel')
            if 법인명_element:
                법인명 = 법인명_element[0].text.strip()
            else:
                print(f"게시글 {i+1} - 법인명 요소를 찾을 수 없습니다.")
                법인명 = ''

            # 문서번호 추출
            문서번호_element = driver.find_elements(By.XPATH, '//th[contains(text(),"문서번호")]/following-sibling::td[1]')
            if 문서번호_element:
                문서번호 = 문서번호_element[0].text.strip()
            else:
                print(f"게시글 {i+1} - 문서번호 요소를 찾을 수 없습니다.")
                문서번호 = ''

            # 제목 추출
            제목_element = driver.find_elements(By.CSS_SELECTOR, 'td.approval_text')
            if 제목_element:
                제목_text = 제목_element[0].text.strip()
                제목 = 제목_text.replace(법인명, '').strip()
            else:
                print(f"게시글 {i+1} - 제목 요소를 찾을 수 없습니다.")
                제목 = ''

            # 합의 담당자 추출
            합의담당자_element = driver.find_elements(By.XPATH, '//th[text()="합의선"]/following::tr[@class="name"][1]/td[@class="td_point"]')
            if 합의담당자_element:
                합의담당자 = 합의담당자_element[0].text.strip()
            else:
                print(f"게시글 {i+1} - 합의 담당자 요소를 찾을 수 없습니다.")
                합의담당자 = ''

            # 링크 추출
            링크 = driver.current_url

            # 데이터 저장
            data = {
                '결재일': 결재일_text,
                '년': 년,
                '월': 월,
                '일': 일,
                '주차': '',          # 빈 문자열 할당
                '법인명': 법인명,
                '문서번호': 문서번호,
                '제목': 제목,
                '업무 유형': '',      # 빈 문자열 할당
                '추출 위치': '',      # 빈 문자열 할당
                '담당 부서': '',      # 빈 문자열 할당
                '신청자': 신청자,
                '합의 담당자': 합의담당자,
                '링크': 링크,
                '진행 구분': ''       # 빈 문자열 할당
            }
            data_list.append(data)

            print(f"게시글 {i+1} - 데이터 추출 완료: {data}")

        except Exception as e:
            print(f"게시글 {i+1} - 데이터 추출 중 오류 발생: {e}")
            traceback.print_exc()

        finally:
            # 창 닫기 및 원래 창으로 전환
            driver.close()
            driver.switch_to.window(driver.window_handles[0])

            # 잠시 대기
            time.sleep(2)

    # 데이터프레임 생성
    df = pd.DataFrame(data_list)
    print(f"총 추출된 게시글 수: {len(df)}")

    ######################################엑셀화##############################################

    if not df.empty:
        try:
            # 기존 엑셀 파일 불러오기
            wb = load_workbook(excel_file)
            if worksheet_name not in wb.sheetnames:
                print(f"워크시트 '{worksheet_name}'이(가) 존재하지 않습니다.")
                driver.quit()
                sys.exit()
            ws = wb[worksheet_name]

            # 'NO' 열에서 데이터가 있는 마지막 행 찾기
            last_row = ws.max_row
            while last_row >= 6:  # 데이터가 시작되는 행 번호는 6
                if ws.cell(row=last_row, column=2).value is not None:
                    break
                last_row -= 1

            # 새로운 데이터 입력 시작 행
            if last_row < 6:
                start_row = 6  # 데이터 시작 행
            else:
                start_row = last_row + 1

            # 다음 'NO' 값 결정
            if last_row >= 6:
                last_no = ws.cell(row=last_row, column=2).value
                if isinstance(last_no, int):
                    next_no = last_no + 1
                else:
                    next_no = 1
            else:
                next_no = 1

            # 데이터프레임의 열 순서 조정 (엑셀의 열 순서와 일치하도록)
            df = df[['결재일', '년', '월', '일', '주차', '법인명', '문서번호', '제목', '업무 유형', '추출 위치', '담당 부서', '신청자', '합의 담당자', '링크', '진행 구분']]

            # 열 매핑 설정 (데이터프레임 열 이름과 엑셀 열 인덱스 매핑)
            column_mapping = {
                '결재일': 3,     # C
                '년': 4,         # D
                '월': 5,         # E
                '일': 6,         # F
                '주차': 7,       # G
                '법인명': 8,     # H
                '문서번호': 9,   # I
                '제목': 10,      # J
                '업무 유형': 11, # K
                '추출 위치': 12, # L
                '담당 부서': 13, # M
                '신청자': 14,    # N
                '합의 담당자': 15, # O
                '링크': 16,      # P
                '진행 구분': 17   # Q
            }

            # 데이터프레임을 엑셀 워크시트에 쓰기
            for idx, row in df.iterrows():
                # 'NO' 값 입력 (열 인덱스 2는 B열)
                ws.cell(row=start_row, column=2, value=next_no)
                next_no += 1

                # 각 열에 데이터 입력
                for col_name, col_idx in column_mapping.items():
                    value = row[col_name]
                    ws.cell(row=start_row, column=col_idx, value=value)
                start_row += 1

            # 엑셀 파일 저장
            wb.save(excel_file)
            print(f"데이터가 성공적으로 '{excel_file}' 파일에 저장되었습니다.")

        except Exception as e:
            print("엑셀 파일 처리 중 오류가 발생했습니다.")
            print(e)
            traceback.print_exc()
    else:
        print("추출된 데이터가 없습니다.")

except Exception as e:
    print("스크립트 실행 중 예상치 못한 오류가 발생했습니다.")
    print(e)
    traceback.print_exc()
finally:
    # 브라우저 종료
    driver.quit()
    print("브라우저가 종료되었습니다.")