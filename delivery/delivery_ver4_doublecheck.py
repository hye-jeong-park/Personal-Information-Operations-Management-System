import os
import re
import sys
import time
import traceback
import getpass
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook

# 크롤링할 게시글 개수 설정 (첫 번째 게시글 제외)
CRAWL_LIMIT = 10

# 엑셀 파일 경로 및 워크시트 이름 설정
excel_file = r'C:\Users\PHJ\output\개인정보 운영대장.xlsx'
worksheet_name = '개인정보 추출 및 이용 관리'

def extract_corporate_name(full_text):
    """
    법인명 추출: "게임사업3본부 K사업팀 / 홍길동님" 중 "게임사업3본부"만 추출
    """
    if '/' in full_text:
        return full_text.split('/')[0].strip().split()[0]
    return full_text.strip().split()[0]

def extract_file_info(file_info):
    """
    파일형식 및 파일 용량 추출
    """
    # 파일명과 용량을 '&' 또는 '('로 분리
    file_match = re.match(r'(.+?)\s*(?:&|[(])\s*([\d,\.]+\s*[KMGT]?B)', file_info, re.IGNORECASE)
    if file_match:
        filename_part = file_match.group(1).strip()
        size_part = file_match.group(2).strip()
    else:
        # '&'나 '('가 없을 경우 전체 문자열에서 용량 부분만 추출
        filename_part = file_info.strip()
        size_match = re.search(r'([\d,\.]+\s*[KMGT]?B)', filename_part, re.IGNORECASE)
        if size_match:
            size_part = size_match.group(1).strip()
            filename_part = filename_part.replace(size_part, '').strip()
        else:
            size_part = ''

    # 파일형식 결정
    if '.zip' in filename_part.lower():
        file_type = 'Zip'
    elif '.xlsx' in filename_part.lower():
        file_type = 'Excel'
    else:
        file_type = ''

    # 파일 용량 추출 (예: "24.5KB")
    size_match = re.match(r'([\d,\.]+)\s*([KMGT]?B)', size_part, re.IGNORECASE)
    if size_match:
        size_numeric = size_match.group(1).replace(',', '')
        size_unit = size_match.group(2).upper()
        file_size = f"{size_numeric} {size_unit}"
    else:
        file_size = size_part

    return file_type, file_size

def find_section_text(driver, section_titles):
    """
    특정 섹션의 제목을 기반으로 해당 섹션의 내용을 추출하는 함수
    section_titles: 섹션 제목의 리스트
    """
    try:
        # 모든 <tr> 요소를 반복하면서 섹션 찾기
        tr_elements = driver.find_elements(By.XPATH, '//table//tr')
        for tr in tr_elements:
            try:
                # 각 <tr>에서 첫 번째 <td> 요소의 텍스트 추출
                td_elements = tr.find_elements(By.TAG_NAME, 'td')
                if len(td_elements) >=2:
                    th_td = td_elements[0]
                    spans = th_td.find_elements(By.TAG_NAME, 'span')
                    header_text = ''.join([span.text.strip() for span in spans])

                    for section_title in section_titles:
                        if section_title in header_text:
                            # 해당 <tr>의 두 번째 <td> 요소의 텍스트 추출
                            value_td = td_elements[1]
                            return value_td.text.strip()
            except Exception as e:
                continue
        return None
    except Exception as e:
        print(f"find_section_text 오류: {e}")
        return None
    
def extract_attachment_info(driver):
    """
    메인 문서 내의 첨부파일 정보를 추출하는 함수
    """
    파일형식 = ''
    파일용량 = ''
    # 먼저 iframe 밖의 'attmRead' 영역에서 파일 정보 추출 시도
    try:
        attm_read_div = driver.find_element(By.ID, 'attmRead')
        print("첨부파일 div 찾음: attmRead")
        # 파일용량 추출
        try:
            file_size_element = attm_read_div.find_element(By.XPATH, './/span[@class="attm-size"]')
            size_text = file_size_element.text.strip()
            size_match = re.match(r'([\d,\.]+)\s*([KMGT]?B)', size_text, re.IGNORECASE)
            if size_match:
                size_numeric = size_match.group(1).replace(',', '')
                size_unit = size_match.group(2).upper()
                파일용량 = f"{size_numeric} {size_unit}"
            else:
                파일용량 = size_text
            print(f"파일용량 추출: {파일용량}")
        except Exception as e:
            print(f"파일용량 추출 중 오류 발생: {e}")

        # 파일형식 추출
        try:
            filename_element = attm_read_div.find_element(By.XPATH, './/ul[contains(@class, "attm-list")]/li/a/strong')
            filename = filename_element.text.strip()
            if '.zip' in filename.lower():
                파일형식 = 'Zip'
            elif '.xlsx' in filename.lower():
                파일형식 = 'Excel'
            else:
                파일형식 = ''
            print(f"파일형식 추출: {파일형식}")
        except Exception as e:
            print(f"파일형식 추출 중 오류 발생: {e}")
            파일형식 = ''
    except Exception as e:
        print(f"attmRead를 찾을 수 없음: {e}")
        # 'attmRead'가 없을 경우
        파일형식 = ''
        파일용량 = ''

    # 파일형식과 파일용량이 없으면 iframe 내에서 추출 시도
    if not 파일형식 and not 파일용량:
        try:
            # iframe으로 전환
            iframe = driver.find_element(By.ID, 'ifa_form')
            driver.switch_to.frame(iframe)
            print("iframe으로 전환하여 파일 정보 추출 시도")
            file_text = find_section_text(driver, ['파밀명 및 용량 (KB)', '파일명 및 용량 (KB)'])
            if file_text:
                print(f"iframe 내에서 파일 정보 추출 시작: {file_text}")
                파일형식, 파일용량 = extract_file_info(file_text)
                print(f"iframe 내에서 파일 정보 추출 완료: {파일형식}, {파일용량}")
            else:
                print("iframe 내에서 파일 정보 섹션을 찾을 수 없습니다.")
            # iframe에서 나옴
            driver.switch_to.default_content()
        except Exception as e:
            print(f"iframe에서 파일 정보 추출 중 오류 발생: {e}")
            driver.switch_to.default_content()
    return 파일형식, 파일용량

def main():
    # 웹드라이버 설정 (옵션 추가 가능)
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    # options.add_argument("--headless")  # 필요 시 헤드리스 모드 활성화
    driver = webdriver.Chrome(options=options)

    try:
        # 로그인 페이지로 이동
        driver.get('https://gw.com2us.com/')
        
        # 로그인 처리
        username_input = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, 'Username'))
        )
        password_input = driver.find_element(By.ID, 'Password')
        
        # 사용자로부터 아이디와 비밀번호 입력받기
        username = input('아이디를 입력하세요: ')
        password = getpass.getpass('비밀번호를 입력하세요: ')
        
        # 아이디와 비밀번호 입력
        username_input.send_keys(username)
        password_input.send_keys(password)
        
        # 로그인 버튼 클릭
        login_button = driver.find_element(By.CLASS_NAME, 'btnLogin')
        login_button.click()
        
        # 로그인 성공 여부 확인
        try:
            WebDriverWait(driver, 20).until(
                EC.url_changes('https://gw.com2us.com/')
            )
            current_url = driver.current_url
            print(f"로그인 후 현재 URL: {current_url}")
            
            if 'login' in current_url.lower():
                print("로그인에 실패하였습니다.")
                driver.quit()
                sys.exit()
        except Exception as e:
            print("로그인 성공 여부를 확인할 수 없습니다.")
            traceback.print_exc()
            driver.quit()
            sys.exit()
        
        # 업무지원 > 개인정보 파일 전송 페이지로 이동
        driver.get('https://gw.com2us.com/emate_app/00001/bbs/b2307140306.nsf/view?readform&viewname=view01')
        
        # 페이지 이동 후 현재 URL 출력
        print(f"페이지 이동 후 현재 URL: {driver.current_url}")
        
        # 게시글 목록 가져오기
        posts = WebDriverWait(driver, 20).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'tr[class*="dhx_skyblue"]'))
        )
        total_posts = len(posts)
        print(f"총 게시글 수: {total_posts}")
        
        if total_posts <= 1:
            print("처리할 게시글이 없습니다. (첫 번째 게시글만 존재)")
            driver.quit()
            sys.exit()
        
        # 크롤링할 게시글 개수 설정 (첫 번째 게시글 제외)
        limit = min(CRAWL_LIMIT, total_posts - 1)
        print(f"크롤링할 게시글 개수: {limit}")
        
        data_list = []
        
        for i in range(1, limit + 1):  # 첫 번째 게시글은 인덱스 0이므로 1부터 시작
            # 게시글 목록을 다시 가져옵니다. (동적 페이지일 경우 필요)
            posts = driver.find_elements(By.CSS_SELECTOR, 'tr[class*="dhx_skyblue"]')
            if i >= len(posts):
                print(f"게시글 {i+1}은 존재하지 않습니다. 종료합니다.")
                break
            post = posts[i]
        
            try:
                # 해당 행의 모든 td 요소를 가져옵니다.
                tds = post.find_elements(By.TAG_NAME, 'td')
        
                # 등록일 추출 (5번째 td, 0-based index)
                if len(tds) >= 5:
                    등록일_td = tds[4]
                    등록일_text = 등록일_td.get_attribute('title').strip() if 등록일_td.get_attribute('title') else 등록일_td.text.strip()
                else:
                    print(f"게시글 {i+1}: 등록일 정보가 부족합니다.")
                    등록일_text = ''

                # 작성자 추출 (3번째 td)
                if len(tds) >= 3:
                    작성자_td = tds[2]
                    # 작성자가 <span> 태그 내에 있는 경우
                    try:
                        작성자 = 작성자_td.find_element(By.TAG_NAME, 'span').text.strip()
                    except:
                        # <span> 태그가 없는 경우
                        작성자 = 작성자_td.text.strip()
                else:
                    print(f"게시글 {i+1}: 작성자 정보가 부족합니다.")
                    작성자 = ''

            except Exception as e:
                print(f"목록에서 데이터 추출 중 오류 발생 (게시글 {i+1}): {e}")
                등록일_text = 작성자 = ''
                continue  # 오류 발생 시 다음 게시글로 이동

            # 요소가 화면에 보이도록 스크롤합니다.
            driver.execute_script("arguments[0].scrollIntoView();", post)

            # 클릭 가능할 때까지 대기합니다.
            try:
                WebDriverWait(driver, 20).until(EC.element_to_be_clickable(post))
            except Exception as e:
                print(f"게시글을 클릭할 수 없습니다 (게시글 {i+1}): {e}")
                traceback.print_exc()
                continue

            # 게시글 클릭하여 상세 페이지 열기
            try:
                post.click()
            except Exception as e:
                print(f"게시글 클릭 중 오류 발생 (게시글 {i+1}): {e}")
                traceback.print_exc()
                continue

            # 새로운 창으로 전환
            try:
                WebDriverWait(driver, 20).until(EC.number_of_windows_to_be(2))
                driver.switch_to.window(driver.window_handles[-1])
                print(f"게시글 {i+1}: 새 창으로 전환")
            except Exception as e:
                print(f"새 창으로 전환 중 오류 발생 (게시글 {i+1}): {e}")
                traceback.print_exc()
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                continue

            # 필요한 시간 대기 (페이지 로딩)
            try:
                WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.ID, 'HeaderTable'))
                )
                print(f"게시글 {i+1}: 상세 페이지 로딩 완료")
            except Exception as e:
                print(f"상세 페이지 로딩 중 오류 발생 (게시글 {i+1}): {e}")
                traceback.print_exc()
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                continue

            try:
                # 현재 창 제목 출력
                print(f"게시글 {i+1}: 현재 창 제목: {driver.title}")
        
                # 제목 추출
                try:
                    제목 = driver.find_element(By.ID, 'DisSubject').text.strip()
                except Exception as e:
                    제목 = ''
                    print(f"게시글 {i+1}: 제목을 찾을 수 없습니다: {e}")
        
                # 작성자 추출
                try:
                    작성자_full = driver.find_element(By.ID, 'DismyName').text.strip()
                except Exception as e:
                    작성자_full = ''
                    print(f"게시글 {i+1}: 작성자 전체 이름을 찾을 수 없습니다: {e}")
        
                # 등록일 추출
                try:
                    등록일_text_detail = driver.find_element(By.ID, 'DiscDate').text.strip()
                except Exception as e:
                    등록일_text_detail = ''
                    print(f"게시글 {i+1}: 등록일을 찾을 수 없습니다: {e}")
        
                # 초기값 설정
                파일형식 = ''
                파일용량 = ''
                법인명 = ''
                개인정보_수 = 0
                진행_구분 = ''
        
                # 첨부파일 정보 추출
                try:
                    파일형식, 파일용량 = extract_attachment_info(driver)
                    print(f"게시글 {i+1}: 첨부파일 정보 추출 완료: {파일형식}, {파일용량}")
                except Exception as e:
                    print(f"게시글 {i+1}: 첨부파일 정보 추출 중 오류 발생: {e}")
                    traceback.print_exc()

                # iframe 내의 '수신자 (부서, 이름)' 및 '추출된 항목 및 건수'는 항상 추출
                try:
                    iframe = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, 'ifa_form'))
                    )
                    driver.switch_to.frame(iframe)
                    print(f"게시글 {i+1}: iframe으로 전환")
                except Exception as e:
                    print(f"게시글 {i+1}: iframe을 찾거나 전환하는 중 오류 발생: {e}")
                    traceback.print_exc()
                # iframe 내에서 데이터 추출
                # 수신자 (부서, 이름) 추출
                try:
                    recipient_text = find_section_text(driver, ['수신자 (부서, 이름)', '수신자(부서, 이름)'])
                    if recipient_text:
                        법인명 = extract_corporate_name(recipient_text)
                        print(f"게시글 {i+1}: 수신자 정보 추출 완료: {법인명}")
                    else:
                        법인명 = ''
                        print(f"게시글 {i+1}: 수신자 정보를 찾을 수 없습니다.")
                except Exception as e:
                    법인명 = ''
                    print(f"게시글 {i+1}: 수신자 정보 추출 중 오류 발생: {e}")
                    traceback.print_exc()

                # 추출된 항목 및 건수 추출
                try:
                    item_text = find_section_text(driver, ['추출된 항목 및 건수'])
                    if item_text:
                        # 모든 "건" 앞의 숫자를 추출하여 합산
                        건수_matches = re.findall(r'(\d{1,3}(?:,\d{3})*)\s*건', item_text)
                        개인정보_수 = sum(int(match.replace(',', '')) for match in 건수_matches)
                        print(f"게시글 {i+1}: 개인정보 수 추출 완료: {개인정보_수}")
                    else:
                        개인정보_수 = 0
                        print(f"게시글 {i+1}: '추출된 항목 및 건수' 섹션을 찾을 수 없습니다.")
                except Exception as e:
                    개인정보_수 = 0
                    print(f"게시글 {i+1}: 개인정보 수 추출 중 오류 발생: {e}")
                    traceback.print_exc()

                # iframe에서 나옴
                driver.switch_to.default_content()
        
                # 진행 구분 설정: '제목'에 '추출완료' 포함 시 "다운 완료"
                if '추출완료' in 제목:
                    진행_구분 = '다운 완료'
                else:
                    진행_구분 = ''
        
                # 데이터 저장
                data = {
                    '등록일': 등록일_text if 등록일_text else 등록일_text_detail,  # 상세 등록일이 있으면 사용
                    '법인명': 법인명,
                    '제목': 제목,
                    '작성자': 작성자_full,
                    '링크': driver.current_url,
                    '파일형식': 파일형식,
                    '파일 용량': 파일용량,
                    '고유식별정보(수)': '',  # 공백으로 저장
                    '개인정보(수)': 개인정보_수,
                    '진행 구분': 진행_구분
                }
                data_list.append(data)
                print(f"게시글 {i+1}: 데이터 추출 완료")
            except Exception as e:
                print(f"게시글 {i+1}: 데이터 추출 중 오류 발생: {e}")
                traceback.print_exc()

            # 창 닫기 및 원래 창으로 전환
            try:
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
            except Exception as e:
                print(f"게시글 {i+1}: 창 닫기 및 전환 중 오류 발생: {e}")
                traceback.print_exc()
                continue

            # 잠시 대기
            time.sleep(2)
        
        # 데이터프레임 생성
        df = pd.DataFrame(data_list)
        
        ######################################엑셀화##############################################
        
        if not df.empty:
            try:
                # 기존 엑셀 파일 불러오기
                wb = load_workbook(excel_file)
                if worksheet_name not in wb.sheetnames:
                    print(f"워크시트 '{worksheet_name}'이(가) 존재하지 않습니다.")
                    driver.quit()
                    sys.exit()
                ws = wb[worksheet_name]
            
                # 'S' 열(등록일)에서 데이터가 있는 마지막 행 찾기
                last_row = ws.max_row
                while last_row >= 5:  # 데이터가 시작되는 행 번호는 5
                    if ws.cell(row=last_row, column=19).value is not None:  # S열 (등록일) 확인
                        break
                    last_row -= 1
            
                # 새로운 데이터 입력 시작 행
                if last_row < 5:
                    start_row = 5  # 데이터 시작 행
                else:
                    start_row = last_row + 1
            
                # 데이터프레임의 열 순서 조정 (엑셀의 열 순서와 일치하도록)
                df = df[['등록일', '법인명', '제목', '작성자', '링크', '파일형식', '파일 용량', '고유식별정보(수)', '개인정보(수)', '진행 구분']]
            
                # 열 매핑 설정 (데이터프레임 열 이름과 엑셀 열 인덱스 매핑)
                column_mapping = {
                    '등록일': 19,           # S
                    '법인명': 20,           # T
                    '제목': 21,             # U
                    '작성자': 22,           # V
                    '링크': 23,             # W
                    '파일형식': 24,         # X
                    '파일 용량': 25,        # Y
                    '고유식별정보(수)': 26, # Z
                    '개인정보(수)': 27,     # AA
                    '진행 구분': 28         # AB
                }
            
                # 데이터프레임을 엑셀 워크시트에 쓰기
                for idx, row in df.iterrows():
                    # 각 열에 데이터 입력
                    for col_name, col_idx in column_mapping.items():
                        value = row[col_name]
                        ws.cell(row=start_row, column=col_idx, value=value)
                    start_row += 1
            
                # 엑셀 파일 저장
                wb.save(excel_file)
                print(f"데이터가 성공적으로 '{excel_file}' 파일에 저장되었습니다.")

            except Exception as e:
                print("엑셀 파일 처리 중 오류가 발생했습니다.")
                print(e)
                traceback.print_exc()
        else:
            print("추출된 데이터가 없습니다.")
    
    except Exception as e:
        print("스크립트 실행 중 예상치 못한 오류가 발생했습니다.")
        print(e)
        traceback.print_exc()
    finally:
        # 브라우저 종료
        driver.quit()
        print("브라우저가 종료되었습니다.")

if __name__ == "__main__":
    main()





###########################################게사글 8,9,10은 모두 잘 나오는 코드##############################################
# import os
# import re
# import sys
# import time
# import traceback
# import getpass
# import pandas as pd
# from selenium import webdriver
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# from openpyxl import load_workbook

# # 크롤링할 게시글 개수 설정 (첫 번째 게시글 제외)
# CRAWL_LIMIT = 10 

# # 엑셀 파일 경로 및 워크시트 이름 설정
# excel_file = r'C:\Users\PHJ\output\개인정보 운영대장.xlsx'
# worksheet_name = '개인정보 추출 및 이용 관리'

# def extract_corporate_name(full_text):
#     """
#     법인명 추출: "게임사업3본부 K사업팀 / 홍길동님" 중 "게임사업3본부"만 추출
#     """
#     if '/' in full_text:
#         return full_text.split('/')[0].strip().split()[0]
#     return full_text.strip().split()[0]

# def extract_file_info(file_info):
#     """
#     파일형식 및 파일 용량 추출
#     """
#     print(f"extract_file_info - 입력 값: {file_info}")
#     # 파일명과 용량을 마지막 괄호 '('로 분리
#     match = re.match(r'(.+?)\s*\(([\d,\.]+\s*[KMGT]?B)\)', file_info, re.IGNORECASE)
#     if match:
#         filename_part = match.group(1).strip()
#         size_part = match.group(2).strip()
#     else:
#         # 괄호가 없는 경우 다른 구분자를 사용하여 분리
#         separators = ['&', ',', ';', ':']
#         parts = [file_info]
#         for sep in separators:
#             if sep in file_info:
#                 parts = file_info.split(sep)
#                 break
#         filename_part = parts[0].strip()
#         size_part = ''
#         for part in parts[1:]:
#             size_match = re.search(r'([\d,\.]+\s*[KMGT]?B)', part, re.IGNORECASE)
#             if size_match:
#                 size_part = size_match.group(1).strip()
#                 break

#     # 파일형식 결정
#     if '.zip' in filename_part.lower():
#         file_type = 'Zip'
#     elif '.xlsx' in filename_part.lower():
#         file_type = 'Excel'
#     else:
#         file_type = ''

#     # 파일 용량 처리
#     if size_part:
#         size_numeric = re.findall(r'[\d,\.]+', size_part)
#         size_unit = re.findall(r'[KMGT]?B', size_part, re.IGNORECASE)
#         if size_numeric and size_unit:
#             size_numeric = size_numeric[0].replace(',', '')
#             size_unit = size_unit[0].upper()
#             file_size = f"{size_numeric} {size_unit}"
#         else:
#             file_size = size_part
#     else:
#         file_size = ''

#     print(f"추출된 파일명: {filename_part}, 추출된 용량: {size_part}")
#     print(f"결과 - 파일형식: {file_type}, 파일용량: {file_size}")
#     return file_type, file_size

# def find_section_text(driver, section_titles):
#     """
#     특정 섹션의 제목을 기반으로 해당 섹션의 내용을 추출하는 함수
#     section_titles: 섹션 제목의 리스트
#     """
#     try:
#         # 모든 <tr> 요소를 반복하면서 섹션 찾기
#         tr_elements = driver.find_elements(By.XPATH, '//table//tr')
#         for tr in tr_elements:
#             try:
#                 # 각 <tr>에서 첫 번째 <td> 요소의 텍스트 추출
#                 td_elements = tr.find_elements(By.TAG_NAME, 'td')
#                 if len(td_elements) >=2:
#                     th_td = td_elements[0]
#                     spans = th_td.find_elements(By.TAG_NAME, 'span')
#                     header_text = ''.join([span.text.strip() for span in spans])

#                     for section_title in section_titles:
#                         if section_title.strip().lower() in header_text.strip().lower():
#                             # 해당 <tr>의 두 번째 <td> 요소의 텍스트 추출
#                             value_td = td_elements[1]
#                             return value_td.text.strip()
#             except Exception as e:
#                 continue
#         return None
#     except Exception as e:
#         print(f"find_section_text 오류: {e}")
#         return None

# def extract_attachment_info(driver):
#     """
#     메인 문서 내의 첨부파일 정보를 추출하는 함수
#     """
#     # 첨부파일 정보를 저장할 변수 초기화
#     file_type = ''
#     file_size = ''

#     # 우선 'attmReadWrap'을 찾는다
#     try:
#         attm_read_div = driver.find_element(By.ID, 'attmReadWrap')
#         print("첨부파일 div 찾음: attmReadWrap")
#         # 첨부파일 목록 찾기
#         attachment_links = attm_read_div.find_elements(By.XPATH, './/ul[contains(@class, "attm-list")]/li/a')
        
#         attachments = []
#         for link in attachment_links:
#             try:
#                 # 파일명과 파일 크기 추출
#                 filename_element = link.find_element(By.TAG_NAME, 'strong')
#                 filename = filename_element.text.strip()
                
#                 size_element = link.find_element(By.XPATH, './/span[contains(@class, "attm-bd")]')
#                 size_text = size_element.text.strip().replace('(', '').replace(')', '')
                
#                 file_info = f"{filename} ({size_text})" if size_text else filename
#                 file_type, file_size = extract_file_info(file_info)
#                 attachments.append((file_type, file_size))
#                 print(f"첨부파일 추출: {filename}, {file_type}, {file_size}")
#             except Exception as e:
#                 print(f"첨부파일 항목 추출 중 오류 발생: {e}")
#                 continue

#         if attachments:
#             # 여러 첨부파일이 있을 경우 첫 번째 파일 정보 사용
#             return attachments[0]
#         else:
#             return '', ''
#     except Exception as e:
#         print(f"attmReadWrap을 찾을 수 없음: {e}")
#         # 'attmReadWrap'을 찾을 수 없을 때 alternative 방법 시도
#         # 'printAttachNot' 섹션에서 첨부파일 정보 추출 시도
#         try:
#             # 'printAttachNot' 섹션 찾기
#             print_attach_tr = driver.find_element(By.ID, 'printAttachNot')
#             print("첨부파일 tr 찾음: printAttachNot")
#             # 'td' 내에서 파일명과 용량 추출
#             td = print_attach_tr.find_element(By.TAG_NAME, 'td')
#             attachment_links = td.find_elements(By.TAG_NAME, 'a')
#             attachments = []
#             for link in attachment_links:
#                 try:
#                     link_text = link.text.strip()
#                     # link_text 예: "파일명.xlsx (28.00 KB)"
#                     file_type, file_size = extract_file_info(link_text)
#                     attachments.append((file_type, file_size))
#                     print(f"첨부파일 추출: {link_text}, {file_type}, {file_size}")
#                 except Exception as e:
#                     print(f"첨부파일 항목 추출 중 오류 발생: {e}")
#                     continue
#             if attachments:
#                 return attachments[0]
#             else:
#                 return '', ''
#         except Exception as e:
#             print(f"첨부파일 정보 추출 실패: {e}")
#             # iframe 내에서 첨부파일 정보 찾기 시도
#             try:
#                 iframe = driver.find_element(By.ID, 'ifa_form')
#                 driver.switch_to.frame(iframe)
#                 print("iframe으로 전환하여 첨부파일 정보 찾기")

#                 # 섹션 제목 리스트 확장
#                 section_titles = [
#                     '파밀명 및 용량 (KB)',
#                     '파일명 및 용량 (KB)',
#                     '첨부파일명 및 용량',
#                     '파일명 및 용량',
#                     '파일명',
#                     '파일 정보',
#                     '첨부파일 정보'
#                 ]

#                 file_text = find_section_text(driver, section_titles)
#                 if file_text:
#                     print(f"추출한 파일 정보 텍스트: {file_text}")
#                     file_type_iframe, file_size_iframe = extract_file_info(file_text)
#                     file_type = file_type_iframe
#                     file_size = file_size_iframe
#                     print(f"iframe 내 첨부파일 정보 추출 완료: {file_type}, {file_size}")
#                 else:
#                     print("iframe 내에서 첨부파일 정보를 찾을 수 없습니다.")
#             except Exception as e:
#                 print(f"iframe에서 첨부파일 정보를 찾는 중 오류 발생: {e}")
#             finally:
#                 # iframe에서 기본 문서로 돌아오기
#                 driver.switch_to.default_content()
#             return file_type, file_size

#     # 최종적으로 첨부파일 정보를 찾지 못한 경우
#     return '', ''

# def main():
#     # 웹드라이버 설정 (옵션 추가 가능)
#     options = webdriver.ChromeOptions()
#     options.add_argument("--start-maximized")
#     # options.add_argument("--headless")  # 필요 시 헤드리스 모드 활성화
#     driver = webdriver.Chrome(options=options)

#     try:
#         # 로그인 페이지로 이동
#         driver.get('https://gw.com2us.com/')
        
#         # 로그인 처리
#         username_input = WebDriverWait(driver, 20).until(
#             EC.presence_of_element_located((By.ID, 'Username'))
#         )
#         password_input = driver.find_element(By.ID, 'Password')
        
#         # 사용자로부터 아이디와 비밀번호 입력받기
#         username = input('아이디를 입력하세요: ')
#         password = getpass.getpass('비밀번호를 입력하세요: ')
        
#         # 아이디와 비밀번호 입력
#         username_input.send_keys(username)
#         password_input.send_keys(password)
        
#         # 로그인 버튼 클릭
#         login_button = driver.find_element(By.CLASS_NAME, 'btnLogin')
#         login_button.click()
        
#         # 로그인 성공 여부 확인
#         try:
#             WebDriverWait(driver, 20).until(
#                 EC.url_changes('https://gw.com2us.com/')
#             )
#             current_url = driver.current_url
#             print(f"로그인 후 현재 URL: {current_url}")
            
#             if 'login' in current_url.lower():
#                 print("로그인에 실패하였습니다.")
#                 driver.quit()
#                 sys.exit()
#         except Exception as e:
#             print("로그인 성공 여부를 확인할 수 없습니다.")
#             traceback.print_exc()
#             driver.quit()
#             sys.exit()
        
#         # 업무지원 > 개인정보 파일 전송 페이지로 이동
#         driver.get('https://gw.com2us.com/emate_app/00001/bbs/b2307140306.nsf/view?readform&viewname=view01')
        
#         # 페이지 이동 후 현재 URL 출력
#         print(f"페이지 이동 후 현재 URL: {driver.current_url}")
        
#         # 게시글 목록 가져오기
#         posts = WebDriverWait(driver, 20).until(
#             EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'tr[class*="dhx_skyblue"]'))
#         )
#         total_posts = len(posts)
#         print(f"총 게시글 수: {total_posts}")
        
#         if total_posts <= 1:
#             print("처리할 게시글이 없습니다. (첫 번째 게시글만 존재)")
#             driver.quit()
#             sys.exit()
        
#         # 크롤링할 게시글 개수 설정 (첫 번째 게시글 제외)
#         limit = min(CRAWL_LIMIT, total_posts - 1)
#         print(f"크롤링할 게시글 개수: {limit}")
        
#         data_list = []
        
#         for i in range(1, limit + 1):  # 첫 번째 게시글은 인덱스 0이므로 1부터 시작
#             # 게시글 목록을 다시 가져옵니다. (동적 페이지일 경우 필요)
#             posts = driver.find_elements(By.CSS_SELECTOR, 'tr[class*="dhx_skyblue"]')
#             if i >= len(posts):
#                 print(f"게시글 {i+1}은 존재하지 않습니다. 종료합니다.")
#                 break
#             post = posts[i]
        
#             try:
#                 # 해당 행의 모든 td 요소를 가져옵니다.
#                 tds = post.find_elements(By.TAG_NAME, 'td')
        
#                 # 등록일 추출 (5번째 td, 0-based index)
#                 if len(tds) >= 5:
#                     등록일_td = tds[4]
#                     등록일_text = 등록일_td.get_attribute('title').strip() if 등록일_td.get_attribute('title') else 등록일_td.text.strip()
#                 else:
#                     print(f"게시글 {i+1}: 등록일 정보가 부족합니다.")
#                     등록일_text = ''

#                 # 작성자 추출 (3번째 td)
#                 if len(tds) >= 3:
#                     작성자_td = tds[2]
#                     # 작성자가 <span> 태그 내에 있는 경우
#                     try:
#                         작성자 = 작성자_td.find_element(By.TAG_NAME, 'span').text.strip()
#                     except:
#                         # <span> 태그가 없는 경우
#                         작성자 = 작성자_td.text.strip()
#                 else:
#                     print(f"게시글 {i+1}: 작성자 정보가 부족합니다.")
#                     작성자 = ''

#             except Exception as e:
#                 print(f"목록에서 데이터 추출 중 오류 발생 (게시글 {i+1}): {e}")
#                 등록일_text = 작성자 = ''
#                 continue  # 오류 발생 시 다음 게시글로 이동

#             # 요소가 화면에 보이도록 스크롤합니다.
#             driver.execute_script("arguments[0].scrollIntoView();", post)

#             # 클릭 가능할 때까지 대기합니다.
#             try:
#                 WebDriverWait(driver, 20).until(EC.element_to_be_clickable(post))
#             except Exception as e:
#                 print(f"게시글을 클릭할 수 없습니다 (게시글 {i+1}): {e}")
#                 traceback.print_exc()
#                 continue

#             # 게시글 클릭하여 상세 페이지 열기
#             try:
#                 post.click()
#             except Exception as e:
#                 print(f"게시글 클릭 중 오류 발생 (게시글 {i+1}): {e}")
#                 traceback.print_exc()
#                 continue

#             # 새로운 창으로 전환
#             try:
#                 WebDriverWait(driver, 20).until(EC.number_of_windows_to_be(2))
#                 driver.switch_to.window(driver.window_handles[-1])
#                 print(f"게시글 {i+1}: 새 창으로 전환")
#             except Exception as e:
#                 print(f"새 창으로 전환 중 오류 발생 (게시글 {i+1}): {e}")
#                 traceback.print_exc()
#                 driver.close()
#                 driver.switch_to.window(driver.window_handles[0])
#                 continue

#             # 필요한 시간 대기 (페이지 로딩)
#             try:
#                 WebDriverWait(driver, 20).until(
#                     EC.presence_of_element_located((By.ID, 'HeaderTable'))
#                 )
#                 print(f"게시글 {i+1}: 상세 페이지 로딩 완료")
#             except Exception as e:
#                 print(f"상세 페이지 로딩 중 오류 발생 (게시글 {i+1}): {e}")
#                 traceback.print_exc()
#                 driver.close()
#                 driver.switch_to.window(driver.window_handles[0])
#                 continue

#             try:
#                 # 현재 창 제목 출력
#                 print(f"게시글 {i+1}: 현재 창 제목: {driver.title}")
        
#                 # 제목 추출
#                 try:
#                     제목 = driver.find_element(By.ID, 'DisSubject').text.strip()
#                 except Exception as e:
#                     제목 = ''
#                     print(f"게시글 {i+1}: 제목을 찾을 수 없습니다: {e}")
        
#                 # 작성자 추출
#                 try:
#                     작성자_full = driver.find_element(By.ID, 'DismyName').text.strip()
#                 except Exception as e:
#                     작성자_full = ''
#                     print(f"게시글 {i+1}: 작성자 전체 이름을 찾을 수 없습니다: {e}")
        
#                 # 등록일 추출
#                 try:
#                     등록일_text_detail = driver.find_element(By.ID, 'DiscDate').text.strip()
#                 except Exception as e:
#                     등록일_text_detail = ''
#                     print(f"게시글 {i+1}: 등록일을 찾을 수 없습니다: {e}")
        
#                 # 초기값 설정
#                 파일형식 = ''
#                 파일용량 = ''
#                 법인명 = ''
#                 개인정보_수 = 0
#                 진행_구분 = ''
        
#                 # 첨부파일 정보 추출
#                 파일형식, 파일용량 = extract_attachment_info(driver)
#                 print(f"게시글 {i+1}: 첨부파일 정보 추출 완료: {파일형식}, {파일용량}")

#                 # iframe 내의 '수신자 (부서, 이름)' 및 '추출된 항목 및 건수'는 항상 추출
#                 try:
#                     iframe = WebDriverWait(driver, 10).until(
#                         EC.presence_of_element_located((By.ID, 'ifa_form'))
#                     )
#                     driver.switch_to.frame(iframe)
#                     print(f"게시글 {i+1}: iframe으로 전환")
#                 except Exception as e:
#                     print(f"게시글 {i+1}: iframe을 찾거나 전환하는 중 오류 발생: {e}")
#                     traceback.print_exc()
#                 # iframe 내에서 데이터 추출
#                 # 수신자 (부서, 이름) 추출
#                 try:
#                     recipient_text = find_section_text(driver, ['수신자 (부서, 이름)', '수신자(부서, 이름)'])
#                     if recipient_text:
#                         법인명 = extract_corporate_name(recipient_text)
#                         print(f"게시글 {i+1}: 수신자 정보 추출 완료: {법인명}")
#                     else:
#                         법인명 = ''
#                         print(f"게시글 {i+1}: 수신자 정보를 찾을 수 없습니다.")
#                 except Exception as e:
#                     법인명 = ''
#                     print(f"게시글 {i+1}: 수신자 정보 추출 중 오류 발생: {e}")
#                     traceback.print_exc()

#                 # 추출된 항목 및 건수 추출
#                 try:
#                     item_text = find_section_text(driver, ['추출된 항목 및 건수'])
#                     if item_text:
#                         # 모든 "건" 앞의 숫자를 추출하여 합산
#                         건수_matches = re.findall(r'(\d{1,3}(?:,\d{3})*)\s*건', item_text)
#                         개인정보_수 = sum(int(match.replace(',', '')) for match in 건수_matches)
#                         print(f"게시글 {i+1}: 개인정보 수 추출 완료: {개인정보_수}")
#                     else:
#                         개인정보_수 = 0
#                         print(f"게시글 {i+1}: '추출된 항목 및 건수' 섹션을 찾을 수 없습니다.")
#                 except Exception as e:
#                     개인정보_수 = 0
#                     print(f"게시글 {i+1}: 개인정보 수 추출 중 오류 발생: {e}")
#                     traceback.print_exc()

#                 # <iframe>에서 기본 문서로 돌아오기
#                 driver.switch_to.default_content()
        
#                 # 진행 구분 설정: '제목'에 '추출완료' 포함 시 "다운 완료"
#                 if '추출완료' in 제목:
#                     진행_구분 = '다운 완료'
#                 else:
#                     진행_구분 = ''
        
#                 # 데이터 저장
#                 data = {
#                     '등록일': 등록일_text if 등록일_text else 등록일_text_detail,  # 상세 등록일이 있으면 사용
#                     '법인명': 법인명,
#                     '제목': 제목,
#                     '작성자': 작성자_full,
#                     '링크': driver.current_url,
#                     '파일형식': 파일형식,
#                     '파일 용량': 파일용량,
#                     '고유식별정보(수)': '',  # 공백으로 저장
#                     '개인정보(수)': 개인정보_수,
#                     '진행 구분': 진행_구분
#                 }
#                 data_list.append(data)
#                 print(f"게시글 {i+1}: 데이터 추출 완료")
#             except Exception as e:
#                 print(f"게시글 {i+1}: 데이터 추출 중 오류 발생: {e}")
#                 traceback.print_exc()

#             # 창 닫기 및 원래 창으로 전환
#             try:
#                 driver.close()
#                 driver.switch_to.window(driver.window_handles[0])
#             except Exception as e:
#                 print(f"게시글 {i+1}: 창 닫기 및 전환 중 오류 발생: {e}")
#                 traceback.print_exc()
#                 continue

#             # 잠시 대기
#             time.sleep(2)
        
#         # 데이터프레임 생성
#         df = pd.DataFrame(data_list)
        
#         ######################################엑셀화##############################################
        
#         if not df.empty:
#             try:
#                 # 기존 엑셀 파일 불러오기
#                 wb = load_workbook(excel_file)
#                 if worksheet_name not in wb.sheetnames:
#                     print(f"워크시트 '{worksheet_name}'이(가) 존재하지 않습니다.")
#                     driver.quit()
#                     sys.exit()
#                 ws = wb[worksheet_name]
            
#                 # 'S' 열(등록일)에서 데이터가 있는 마지막 행 찾기
#                 last_row = ws.max_row
#                 while last_row >= 5:  # 데이터가 시작되는 행 번호는 5
#                     if ws.cell(row=last_row, column=19).value is not None:  # S열 (등록일) 확인
#                         break
#                     last_row -= 1
            
#                 # 새로운 데이터 입력 시작 행
#                 if last_row < 5:
#                     start_row = 5  # 데이터 시작 행
#                 else:
#                     start_row = last_row + 1
            
#                 # 데이터프레임의 열 순서 조정 (엑셀의 열 순서와 일치하도록)
#                 df = df[['등록일', '법인명', '제목', '작성자', '링크', '파일형식', '파일 용량', '고유식별정보(수)', '개인정보(수)', '진행 구분']]
            
#                 # 열 매핑 설정 (데이터프레임 열 이름과 엑셀 열 인덱스 매핑)
#                 column_mapping = {
#                     '등록일': 19,           # S
#                     '법인명': 20,           # T
#                     '제목': 21,             # U
#                     '작성자': 22,           # V
#                     '링크': 23,             # W
#                     '파일형식': 24,         # X
#                     '파일 용량': 25,        # Y
#                     '고유식별정보(수)': 26, # Z
#                     '개인정보(수)': 27,     # AA
#                     '진행 구분': 28         # AB
#                 }
            
#                 # 데이터프레임을 엑셀 워크시트에 쓰기
#                 for idx, row in df.iterrows():
#                     # 각 열에 데이터 입력
#                     for col_name, col_idx in column_mapping.items():
#                         value = row[col_name]
#                         ws.cell(row=start_row, column=col_idx, value=value)
#                     start_row += 1
            
#                 # 엑셀 파일 저장
#                 wb.save(excel_file)
#                 print(f"데이터가 성공적으로 '{excel_file}' 파일에 저장되었습니다.")

#             except Exception as e:
#                 print("엑셀 파일 처리 중 오류가 발생했습니다.")
#                 print(e)
#                 traceback.print_exc()
#         else:
#             print("추출된 데이터가 없습니다.")
    
#     except Exception as e:
#         print("스크립트 실행 중 예상치 못한 오류가 발생했습니다.")
#         print(e)
#         traceback.print_exc()
#     finally:
#         # 브라우저 종료
#         driver.quit()
#         print("브라우저가 종료되었습니다.")

# if __name__ == "__main__":
#     main()












###########################################파일형식은 모두 잘 나오나 파일 용량은 게시글8에 대해서만 잘 나오는 코드##############################################
# import os
# import re
# import sys
# import time
# import traceback
# import getpass
# import pandas as pd
# from selenium import webdriver
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# from openpyxl import load_workbook

# # 크롤링할 게시글 개수 설정 (첫 번째 게시글 제외)
# CRAWL_LIMIT = 10 

# # 엑셀 파일 경로 및 워크시트 이름 설정
# excel_file = r'C:\Users\PHJ\output\개인정보 운영대장.xlsx'
# worksheet_name = '개인정보 추출 및 이용 관리'

# def extract_corporate_name(full_text):
#     """
#     법인명 추출: "게임사업3본부 K사업팀 / 홍길동님" 중 "게임사업3본부"만 추출
#     """
#     if '/' in full_text:
#         return full_text.split('/')[0].strip().split()[0]
#     return full_text.strip().split()[0]

# def extract_file_info(file_info):
#     """
#     파일형식 및 파일 용량 추출
#     """
#     # 파일명과 용량을 '('와 ')'로 분리
#     file_match = re.match(r'(.+?)\s*\(([\d,\.]+\s*[KMGT]?B)\)', file_info, re.IGNORECASE)
#     if file_match:
#         filename_part = file_match.group(1).strip()
#         size_part = file_match.group(2).strip()

#         # 파일형식 결정
#         if '.zip' in filename_part.lower():
#             file_type = 'Zip'
#         elif '.xlsx' in filename_part.lower():
#             file_type = 'Excel'
#         else:
#             file_type = ''

#         # 파일 용량 추출 (예: "24.5KB")
#         size_match = re.match(r'([\d,\.]+)\s*([KMGT]?B)', size_part, re.IGNORECASE)
#         if size_match:
#             size_numeric = size_match.group(1).replace(',', '')
#             size_unit = size_match.group(2).upper()
#             file_size = f"{size_numeric} {size_unit}"
#         else:
#             file_size = size_part

#         return file_type, file_size
#     else:
#         # 용량 정보가 없을 경우 파일명만으로 파일형식 결정
#         filename_part = file_info.strip()
#         if '.zip' in filename_part.lower():
#             file_type = 'Zip'
#         elif '.xlsx' in filename_part.lower():
#             file_type = 'Excel'
#         else:
#             file_type = ''
#         file_size = ''
#         return file_type, file_size

# def find_section_text(driver, section_titles):
#     """
#     특정 섹션의 제목을 기반으로 해당 섹션의 내용을 추출하는 함수
#     section_titles: 섹션 제목의 리스트
#     """
#     try:
#         # 모든 <tr> 요소를 반복하면서 섹션 찾기
#         tr_elements = driver.find_elements(By.XPATH, '//table//tr')
#         for tr in tr_elements:
#             try:
#                 # 각 <tr>에서 첫 번째 <td> 요소의 텍스트 추출
#                 td_elements = tr.find_elements(By.TAG_NAME, 'td')
#                 if len(td_elements) >=2:
#                     th_td = td_elements[0]
#                     spans = th_td.find_elements(By.TAG_NAME, 'span')
#                     header_text = ''.join([span.text.strip() for span in spans])

#                     for section_title in section_titles:
#                         if section_title in header_text:
#                             # 해당 <tr>의 두 번째 <td> 요소의 텍스트 추출
#                             value_td = td_elements[1]
#                             return value_td.text.strip()
#             except Exception as e:
#                 continue
#         return None
#     except Exception as e:
#         print(f"find_section_text 오류: {e}")
#         return None

# def extract_attachment_info(driver):
#     """
#     메인 문서 내의 첨부파일 정보를 추출하는 함수
#     """
#     # 우선 'attmReadWrap'을 찾는다
#     try:
#         attm_read_div = driver.find_element(By.ID, 'attmReadWrap')
#         print("첨부파일 div 찾음: attmReadWrap")
#         # 첨부파일 목록 찾기
#         attachment_links = attm_read_div.find_elements(By.XPATH, './/ul[contains(@class, "attm-list")]/li/a')
        
#         attachments = []
#         for link in attachment_links:
#             try:
#                 # 파일명과 파일 크기 추출
#                 filename_element = link.find_element(By.TAG_NAME, 'strong')
#                 filename = filename_element.text.strip()
                
#                 size_element = link.find_element(By.XPATH, './/span[contains(@class, "attm-bd")]')
#                 size_text = size_element.text.strip().replace('(', '').replace(')', '')
                
#                 file_info = f"{filename} ({size_text})" if size_text else filename
#                 file_type, file_size = extract_file_info(file_info)
#                 attachments.append((file_type, file_size))
#                 print(f"첨부파일 추출: {filename}, {file_type}, {file_size}")
#             except Exception as e:
#                 print(f"첨부파일 항목 추출 중 오류 발생: {e}")
#                 continue

#         if attachments:
#             # 여러 첨부파일이 있을 경우 첫 번째 파일 정보 사용
#             return attachments[0]
#         else:
#             return '', ''
#     except Exception as e:
#         print(f"attmReadWrap을 찾을 수 없음: {e}")
#         # 'attmReadWrap'을 찾을 수 없을 때 alternative 방법 시도
#         # 'printAttachNot' 섹션에서 첨부파일 정보 추출 시도
#         try:
#             # 'printAttachNot' 섹션 찾기
#             print_attach_tr = driver.find_element(By.ID, 'printAttachNot')
#             print("첨부파일 tr 찾음: printAttachNot")
#             # 'td' 내에서 파일명과 용량 추출
#             td = print_attach_tr.find_element(By.TAG_NAME, 'td')
#             attachment_links = td.find_elements(By.TAG_NAME, 'a')
#             attachments = []
#             for link in attachment_links:
#                 try:
#                     link_text = link.text.strip()
#                     # link_text 예: "제노니아-제우스 2nd 캠핑키트 대상자 정보 추출 요청의 건.xlsx (28.00 KB)"
#                     file_type, file_size = extract_file_info(link_text)
#                     attachments.append((file_type, file_size))
#                     print(f"첨부파일 추출: {link_text}, {file_type}, {file_size}")
#                 except Exception as e:
#                     print(f"첨부파일 항목 추출 중 오류 발생: {e}")
#                     continue
#             if attachments:
#                 return attachments[0]
#             else:
#                 return '', ''
#         except Exception as e:
#             print(f"첨부파일 정보 추출 실패: {e}")
#             return '', ''

# def main():
#     # 웹드라이버 설정 (옵션 추가 가능)
#     options = webdriver.ChromeOptions()
#     options.add_argument("--start-maximized")
#     # options.add_argument("--headless")  # 필요 시 헤드리스 모드 활성화
#     driver = webdriver.Chrome(options=options)

#     try:
#         # 로그인 페이지로 이동
#         driver.get('https://gw.com2us.com/')
        
#         # 로그인 처리
#         username_input = WebDriverWait(driver, 20).until(
#             EC.presence_of_element_located((By.ID, 'Username'))
#         )
#         password_input = driver.find_element(By.ID, 'Password')
        
#         # 사용자로부터 아이디와 비밀번호 입력받기
#         username = input('아이디를 입력하세요: ')
#         password = getpass.getpass('비밀번호를 입력하세요: ')
        
#         # 아이디와 비밀번호 입력
#         username_input.send_keys(username)
#         password_input.send_keys(password)
        
#         # 로그인 버튼 클릭
#         login_button = driver.find_element(By.CLASS_NAME, 'btnLogin')
#         login_button.click()
        
#         # 로그인 성공 여부 확인
#         try:
#             WebDriverWait(driver, 20).until(
#                 EC.url_changes('https://gw.com2us.com/')
#             )
#             current_url = driver.current_url
#             print(f"로그인 후 현재 URL: {current_url}")
            
#             if 'login' in current_url.lower():
#                 print("로그인에 실패하였습니다.")
#                 driver.quit()
#                 sys.exit()
#         except Exception as e:
#             print("로그인 성공 여부를 확인할 수 없습니다.")
#             traceback.print_exc()
#             driver.quit()
#             sys.exit()
        
#         # 업무지원 > 개인정보 파일 전송 페이지로 이동
#         driver.get('https://gw.com2us.com/emate_app/00001/bbs/b2307140306.nsf/view?readform&viewname=view01')
        
#         # 페이지 이동 후 현재 URL 출력
#         print(f"페이지 이동 후 현재 URL: {driver.current_url}")
        
#         # 게시글 목록 가져오기
#         posts = WebDriverWait(driver, 20).until(
#             EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'tr[class*="dhx_skyblue"]'))
#         )
#         total_posts = len(posts)
#         print(f"총 게시글 수: {total_posts}")
        
#         if total_posts <= 1:
#             print("처리할 게시글이 없습니다. (첫 번째 게시글만 존재)")
#             driver.quit()
#             sys.exit()
        
#         # 크롤링할 게시글 개수 설정 (첫 번째 게시글 제외)
#         limit = min(CRAWL_LIMIT, total_posts - 1)
#         print(f"크롤링할 게시글 개수: {limit}")
        
#         data_list = []
        
#         for i in range(1, limit + 1):  # 첫 번째 게시글은 인덱스 0이므로 1부터 시작
#             # 게시글 목록을 다시 가져옵니다. (동적 페이지일 경우 필요)
#             posts = driver.find_elements(By.CSS_SELECTOR, 'tr[class*="dhx_skyblue"]')
#             if i >= len(posts):
#                 print(f"게시글 {i+1}은 존재하지 않습니다. 종료합니다.")
#                 break
#             post = posts[i]
        
#             try:
#                 # 해당 행의 모든 td 요소를 가져옵니다.
#                 tds = post.find_elements(By.TAG_NAME, 'td')
        
#                 # 등록일 추출 (5번째 td, 0-based index)
#                 if len(tds) >= 5:
#                     등록일_td = tds[4]
#                     등록일_text = 등록일_td.get_attribute('title').strip() if 등록일_td.get_attribute('title') else 등록일_td.text.strip()
#                 else:
#                     print(f"게시글 {i+1}: 등록일 정보가 부족합니다.")
#                     등록일_text = ''

#                 # 작성자 추출 (3번째 td)
#                 if len(tds) >= 3:
#                     작성자_td = tds[2]
#                     # 작성자가 <span> 태그 내에 있는 경우
#                     try:
#                         작성자 = 작성자_td.find_element(By.TAG_NAME, 'span').text.strip()
#                     except:
#                         # <span> 태그가 없는 경우
#                         작성자 = 작성자_td.text.strip()
#                 else:
#                     print(f"게시글 {i+1}: 작성자 정보가 부족합니다.")
#                     작성자 = ''

#             except Exception as e:
#                 print(f"목록에서 데이터 추출 중 오류 발생 (게시글 {i+1}): {e}")
#                 등록일_text = 작성자 = ''
#                 continue  # 오류 발생 시 다음 게시글로 이동

#             # 요소가 화면에 보이도록 스크롤합니다.
#             driver.execute_script("arguments[0].scrollIntoView();", post)

#             # 클릭 가능할 때까지 대기합니다.
#             try:
#                 WebDriverWait(driver, 20).until(EC.element_to_be_clickable(post))
#             except Exception as e:
#                 print(f"게시글을 클릭할 수 없습니다 (게시글 {i+1}): {e}")
#                 traceback.print_exc()
#                 continue

#             # 게시글 클릭하여 상세 페이지 열기
#             try:
#                 post.click()
#             except Exception as e:
#                 print(f"게시글 클릭 중 오류 발생 (게시글 {i+1}): {e}")
#                 traceback.print_exc()
#                 continue

#             # 새로운 창으로 전환
#             try:
#                 WebDriverWait(driver, 20).until(EC.number_of_windows_to_be(2))
#                 driver.switch_to.window(driver.window_handles[-1])
#                 print(f"게시글 {i+1}: 새 창으로 전환")
#             except Exception as e:
#                 print(f"새 창으로 전환 중 오류 발생 (게시글 {i+1}): {e}")
#                 traceback.print_exc()
#                 driver.close()
#                 driver.switch_to.window(driver.window_handles[0])
#                 continue

#             # 필요한 시간 대기 (페이지 로딩)
#             try:
#                 WebDriverWait(driver, 20).until(
#                     EC.presence_of_element_located((By.ID, 'HeaderTable'))
#                 )
#                 print(f"게시글 {i+1}: 상세 페이지 로딩 완료")
#             except Exception as e:
#                 print(f"상세 페이지 로딩 중 오류 발생 (게시글 {i+1}): {e}")
#                 traceback.print_exc()
#                 driver.close()
#                 driver.switch_to.window(driver.window_handles[0])
#                 continue

#             try:
#                 # 현재 창 제목 출력
#                 print(f"게시글 {i+1}: 현재 창 제목: {driver.title}")
        
#                 # 제목 추출
#                 try:
#                     제목 = driver.find_element(By.ID, 'DisSubject').text.strip()
#                 except Exception as e:
#                     제목 = ''
#                     print(f"게시글 {i+1}: 제목을 찾을 수 없습니다: {e}")
        
#                 # 작성자 추출
#                 try:
#                     작성자_full = driver.find_element(By.ID, 'DismyName').text.strip()
#                 except Exception as e:
#                     작성자_full = ''
#                     print(f"게시글 {i+1}: 작성자 전체 이름을 찾을 수 없습니다: {e}")
        
#                 # 등록일 추출
#                 try:
#                     등록일_text_detail = driver.find_element(By.ID, 'DiscDate').text.strip()
#                 except Exception as e:
#                     등록일_text_detail = ''
#                     print(f"게시글 {i+1}: 등록일을 찾을 수 없습니다: {e}")
        
#                 # 초기값 설정
#                 파일형식 = ''
#                 파일용량 = ''
#                 법인명 = ''
#                 개인정보_수 = 0
#                 진행_구분 = ''
        
#                 # 먼저 외부에서 첨부파일 정보 추출 시도
#                 try:
#                     파일형식, 파일용량 = extract_attachment_info(driver)
#                     print(f"게시글 {i+1}: 외부 첨부파일 정보 추출 완료: {파일형식}, {파일용량}")
#                 except Exception as e:
#                     print(f"게시글 {i+1}: 외부 첨부파일 정보 추출 중 오류 발생: {e}")
#                     traceback.print_exc()

#                 # iframe 내의 '수신자 (부서, 이름)' 및 '추출된 항목 및 건수'는 항상 추출
#                 try:
#                     iframe = WebDriverWait(driver, 10).until(
#                         EC.presence_of_element_located((By.ID, 'ifa_form'))
#                     )
#                     driver.switch_to.frame(iframe)
#                     print(f"게시글 {i+1}: iframe으로 전환")
#                 except Exception as e:
#                     print(f"게시글 {i+1}: iframe을 찾거나 전환하는 중 오류 발생: {e}")
#                     traceback.print_exc()
#                 # iframe 내에서 데이터 추출
#                 # 수신자 (부서, 이름) 추출
#                 try:
#                     recipient_text = find_section_text(driver, ['수신자 (부서, 이름)', '수신자(부서, 이름)'])
#                     if recipient_text:
#                         법인명 = extract_corporate_name(recipient_text)
#                         print(f"게시글 {i+1}: 수신자 정보 추출 완료: {법인명}")
#                     else:
#                         법인명 = ''
#                         print(f"게시글 {i+1}: 수신자 정보를 찾을 수 없습니다.")
#                 except Exception as e:
#                     법인명 = ''
#                     print(f"게시글 {i+1}: 수신자 정보 추출 중 오류 발생: {e}")
#                     traceback.print_exc()

#                 # 추출된 항목 및 건수 추출
#                 try:
#                     item_text = find_section_text(driver, ['추출된 항목 및 건수'])
#                     if item_text:
#                         # 모든 "건" 앞의 숫자를 추출하여 합산
#                         건수_matches = re.findall(r'(\d{1,3}(?:,\d{3})*)\s*건', item_text)
#                         개인정보_수 = sum(int(match.replace(',', '')) for match in 건수_matches)
#                         print(f"게시글 {i+1}: 개인정보 수 추출 완료: {개인정보_수}")
#                     else:
#                         개인정보_수 = 0
#                         print(f"게시글 {i+1}: '추출된 항목 및 건수' 섹션을 찾을 수 없습니다.")
#                 except Exception as e:
#                     개인정보_수 = 0
#                     print(f"게시글 {i+1}: 개인정보 수 추출 중 오류 발생: {e}")
#                     traceback.print_exc()

#                 # '파일명 및 용량 (KB)' 또는 '파밀명 및 용량 (KB)' 데이터가 없을 경우 iframe 내에서 추가 추출
#                 if not 파일형식 and not 파일용량:
#                     try:
#                         file_text = find_section_text(driver, ['파밀명 및 용량 (KB)', '파일명 및 용량 (KB)'])
#                         if file_text:
#                             파일형식_iframe, 파일용량_iframe = extract_file_info(file_text)
#                             파일형식 = 파일형식_iframe
#                             파일용량 = 파일용량_iframe
#                             print(f"게시글 {i+1}: iframe 내 '파일명 및 용량 (KB)' 추출 완료: {파일형식}, {파일용량}")
#                         else:
#                             print(f"게시글 {i+1}: '파일명 및 용량 (KB)' 섹션을 찾을 수 없습니다.")
#                     except Exception as e:
#                         파일형식 = ''
#                         파일용량 = ''
#                         print(f"게시글 {i+1}: '파일명 및 용량 (KB)' 추출 중 오류 발생: {e}")
#                         traceback.print_exc()

#                 # <iframe>에서 기본 문서로 돌아오기
#                 driver.switch_to.default_content()
        
#                 # 진행 구분 설정: '제목'에 '추출완료' 포함 시 "다운 완료"
#                 if '추출완료' in 제목:
#                     진행_구분 = '다운 완료'
#                 else:
#                     진행_구분 = ''
        
#                 # 데이터 저장
#                 data = {
#                     '등록일': 등록일_text if 등록일_text else 등록일_text_detail,  # 상세 등록일이 있으면 사용
#                     '법인명': 법인명,
#                     '제목': 제목,
#                     '작성자': 작성자_full,
#                     '링크': driver.current_url,
#                     '파일형식': 파일형식,
#                     '파일 용량': 파일용량,
#                     '고유식별정보(수)': '',  # 공백으로 저장
#                     '개인정보(수)': 개인정보_수,
#                     '진행 구분': 진행_구분
#                 }
#                 data_list.append(data)
#                 print(f"게시글 {i+1}: 데이터 추출 완료")
#             except Exception as e:
#                 print(f"게시글 {i+1}: 데이터 추출 중 오류 발생: {e}")
#                 traceback.print_exc()

#             # 창 닫기 및 원래 창으로 전환
#             try:
#                 driver.close()
#                 driver.switch_to.window(driver.window_handles[0])
#             except Exception as e:
#                 print(f"게시글 {i+1}: 창 닫기 및 전환 중 오류 발생: {e}")
#                 traceback.print_exc()
#                 continue

#             # 잠시 대기
#             time.sleep(2)
        
#         # 데이터프레임 생성
#         df = pd.DataFrame(data_list)
        
#         ######################################엑셀화##############################################
        
#         if not df.empty:
#             try:
#                 # 기존 엑셀 파일 불러오기
#                 wb = load_workbook(excel_file)
#                 if worksheet_name not in wb.sheetnames:
#                     print(f"워크시트 '{worksheet_name}'이(가) 존재하지 않습니다.")
#                     driver.quit()
#                     sys.exit()
#                 ws = wb[worksheet_name]
            
#                 # 'S' 열(등록일)에서 데이터가 있는 마지막 행 찾기
#                 last_row = ws.max_row
#                 while last_row >= 5:  # 데이터가 시작되는 행 번호는 5
#                     if ws.cell(row=last_row, column=19).value is not None:  # S열 (등록일) 확인
#                         break
#                     last_row -= 1
            
#                 # 새로운 데이터 입력 시작 행
#                 if last_row < 5:
#                     start_row = 5  # 데이터 시작 행
#                 else:
#                     start_row = last_row + 1
            
#                 # 데이터프레임의 열 순서 조정 (엑셀의 열 순서와 일치하도록)
#                 df = df[['등록일', '법인명', '제목', '작성자', '링크', '파일형식', '파일 용량', '고유식별정보(수)', '개인정보(수)', '진행 구분']]
            
#                 # 열 매핑 설정 (데이터프레임 열 이름과 엑셀 열 인덱스 매핑)
#                 column_mapping = {
#                     '등록일': 19,           # S
#                     '법인명': 20,           # T
#                     '제목': 21,             # U
#                     '작성자': 22,           # V
#                     '링크': 23,             # W
#                     '파일형식': 24,         # X
#                     '파일 용량': 25,        # Y
#                     '고유식별정보(수)': 26, # Z
#                     '개인정보(수)': 27,     # AA
#                     '진행 구분': 28         # AB
#                 }
            
#                 # 데이터프레임을 엑셀 워크시트에 쓰기
#                 for idx, row in df.iterrows():
#                     # 각 열에 데이터 입력
#                     for col_name, col_idx in column_mapping.items():
#                         value = row[col_name]
#                         ws.cell(row=start_row, column=col_idx, value=value)
#                     start_row += 1
            
#                 # 엑셀 파일 저장
#                 wb.save(excel_file)
#                 print(f"데이터가 성공적으로 '{excel_file}' 파일에 저장되었습니다.")

#             except Exception as e:
#                 print("엑셀 파일 처리 중 오류가 발생했습니다.")
#                 print(e)
#                 traceback.print_exc()
#         else:
#             print("추출된 데이터가 없습니다.")
    
#     except Exception as e:
#         print("스크립트 실행 중 예상치 못한 오류가 발생했습니다.")
#         print(e)
#         traceback.print_exc()
#     finally:
#         # 브라우저 종료
#         driver.quit()
#         print("브라우저가 종료되었습니다.")

# if __name__ == "__main__":
#     main()
