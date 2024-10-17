from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import time
import traceback
import getpass
from openpyxl import load_workbook
import sys
import re

# 크롤링할 게시글 개수 설정 (첫 번째 게시글 제외)
CRAWL_LIMIT = 10 

# 엑셀 파일 경로 및 워크시트 이름 설정
excel_file = r'C:\Users\PHJ\output\개인정보 운영대장.xlsx'
worksheet_name = '개인정보 추출 및 이용 관리'

def extract_corporate_name(full_text):
    # 법인명 추출: "컴투스 운영지원, 홍길동" 중 "컴투스"만 추출
    if ',' in full_text:
        return full_text.split(',')[0].split()[0]
    return full_text.split()[0]

def extract_file_info(file_info):
    # 파일형식 및 파일 용량 추출
    # 예시: "(Confidential)_20241017_103738_smon_lms_target_list.zip (221KB)"
    file_match = re.match(r'\(.*?\)_(.*?)\.(zip|xlsx).*?\((\d+KB)\)', file_info)
    if file_match:
        filename = file_match.group(1) + '.' + file_match.group(2)
        file_size = file_match.group(3)
        if filename.endswith('.zip'):
            file_type = 'Zip'
        elif filename.endswith('.xlsx'):
            file_type = 'Excel'
        else:
            file_type = ''
        return file_type, file_size
    else:
        # 다른 형식의 파일명이 있을 경우
        p_tags = file_info.split('\n')
        if len(p_tags) >= 2:
            filename = p_tags[0].strip()
            file_size = p_tags[1].strip()
            if filename.endswith('.zip'):
                file_type = 'Zip'
            elif filename.endswith('.xlsx'):
                file_type = 'Excel'
            else:
                file_type = ''
            return file_type, file_size
    return '', ''

def main():
    # 웹드라이버 설정
    driver = webdriver.Chrome()

    try:
        # 로그인 페이지로 이동
        driver.get('https://gw.com2us.com/')
        
        # 로그인 처리
        username_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, 'Username'))
        )
        password_input = driver.find_element(By.ID, 'Password')
        
        # 사용자로부터 아이디와 비밀번호 입력받기
        username = input('아이디를 입력하세요: ')
        password = getpass.getpass('비밀번호를 입력하세요: ')
        
        # 아이디와 비밀번호 입력
        username_input.send_keys(username)
        password_input.send_keys(password)
        
        # 로그인 버튼 클릭
        login_button = driver.find_element(By.CLASS_NAME, 'btnLogin')
        login_button.click()
        
        # 로그인 성공 여부 확인
        WebDriverWait(driver, 10).until(
            EC.url_changes('https://gw.com2us.com/')
        )
        current_url = driver.current_url
        print(f"로그인 후 현재 URL: {current_url}")
        
        if 'login' in current_url.lower():
            print("로그인에 실패하였습니다.")
            driver.quit()
            sys.exit()
        
        # 업무지원 > 개인정보 파일 전송 페이지로 이동
        driver.get('https://gw.com2us.com/emate_app/00001/bbs/b2307140306.nsf/view?readform&viewname=view01')
        
        # 페이지 이동 후 현재 URL 출력
        print(f"페이지 이동 후 현재 URL: {driver.current_url}")
        
        # 게시글 목록 가져오기
        posts = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'tr[class*="dhx_skyblue"]'))
        )
        total_posts = len(posts)
        print(f"총 게시글 수: {total_posts}")
        
        if total_posts <= 1:
            print("처리할 게시글이 없습니다. (첫 번째 게시글만 존재)")
            driver.quit()
            sys.exit()
        
        # 크롤링할 게시글 개수 설정 (첫 번째 게시글 제외)
        limit = min(CRAWL_LIMIT, total_posts - 1)
        print(f"크롤링할 게시글 개수: {limit}")
        
        data_list = []
        
        for i in range(1, limit + 1):  # 첫 번째 게시글은 인덱스 0이므로 1부터 시작
            # 게시글 목록을 다시 가져옵니다. (동적 페이지일 경우 필요)
            posts = driver.find_elements(By.CSS_SELECTOR, 'tr[class*="dhx_skyblue"]')
            if i >= len(posts):
                print(f"게시글 {i}은 존재하지 않습니다. 종료합니다.")
                break
            post = posts[i]
        
            try:
                # 해당 행의 모든 td 요소를 가져옵니다.
                tds = post.find_elements(By.TAG_NAME, 'td')
        
                # 등록일 추출 (5번째 td, 0-based index)
                if len(tds) >= 5:
                    등록일_td = tds[4]
                    등록일_text = 등록일_td.get_attribute('title').strip() if 등록일_td.get_attribute('title') else 등록일_td.text.strip()
                else:
                    print(f"게시글 {i}: 등록일 정보가 부족합니다.")
                    등록일_text = ''
        
                # 작성자 추출 (3번째 td)
                if len(tds) >= 3:
                    작성자_td = tds[2]
                    작성자 = 작성자_td.find_element(By.TAG_NAME, 'span').text.strip()
                else:
                    print(f"게시글 {i}: 작성자 정보가 부족합니다.")
                    작성자 = ''
        
            except Exception as e:
                print(f"목록에서 데이터 추출 중 오류 발생 (게시글 {i}): {e}")
                등록일_text = 작성자 = ''
                continue  # 오류 발생 시 다음 게시글로 이동
        
            # 요소가 화면에 보이도록 스크롤합니다.
            driver.execute_script("arguments[0].scrollIntoView();", post)
        
            # 클릭 가능할 때까지 대기합니다.
            try:
                WebDriverWait(driver, 20).until(EC.element_to_be_clickable(post))
            except Exception as e:
                print(f"게시글을 클릭할 수 없습니다 (게시글 {i}): {e}")
                traceback.print_exc()
                continue
    
            # 게시글 클릭하여 상세 페이지 열기
            try:
                post.click()
            except Exception as e:
                print(f"게시글 클릭 중 오류 발생 (게시글 {i}): {e}")
                traceback.print_exc()
                continue
    
            # 새로운 창으로 전환
            try:
                WebDriverWait(driver, 20).until(EC.number_of_windows_to_be(2))
                driver.switch_to.window(driver.window_handles[-1])
            except Exception as e:
                print(f"새 창으로 전환 중 오류 발생 (게시글 {i}): {e}")
                traceback.print_exc()
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                continue
    
            # 필요한 시간 대기 (페이지 로딩)
            try:
                WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.ID, 'HeaderTable'))
                )
                print(f"게시글 {i}: 상세 페이지 로딩 완료")
            except Exception as e:
                print(f"상세 페이지 로딩 중 오류 발생 (게시글 {i}): {e}")
                traceback.print_exc()
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                continue
    
            try:
                # 현재 창 제목 출력
                print(f"게시글 {i}: 현재 창 제목: {driver.title}")
        
                # 제목 추출
                try:
                    제목 = driver.find_element(By.ID, 'DisSubject').text.strip()
                except:
                    제목 = ''
                    print(f"게시글 {i}: 제목을 찾을 수 없습니다.")
        
                # 작성자 추출
                try:
                    작성자_full = driver.find_element(By.ID, 'DismyName').text.strip()
                except:
                    작성자_full = ''
                    print(f"게시글 {i}: 작성자 전체 이름을 찾을 수 없습니다.")
        
                # 등록일 추출
                try:
                    등록일_text = driver.find_element(By.ID, 'DiscDate').text.strip()
                except:
                    등록일_text = ''
                    print(f"게시글 {i}: 등록일을 찾을 수 없습니다.")
        
                # 첨부파일 정보 추출
                파일형식 = ''
                파일용량 = ''
                파일_리스트 = []
                try:
                    attm_list = driver.find_elements(By.CSS_SELECTOR, 'ul.attm-list._r li.item a')
                    for attm in attm_list:
                        href = attm.get_attribute('onclick')
                        # href 예시: downloadFile('URL', 'filename', 'size')
                        match = re.search(r"downloadFile\('(.+?)',\s*'(.+?)',\s*'(.+?)'\)", href)
                        if match:
                            file_url = match.group(1)
                            filename = match.group(2)
                            filesize = match.group(3)
                            if filename.endswith('.zip'):
                                file_type = 'Zip'
                            elif filename.endswith('.xlsx'):
                                file_type = 'Excel'
                            else:
                                file_type = ''
                            파일_리스트.append({
                                '파일형식': file_type,
                                '파일용량': filesize,
                                '링크': file_url,
                                '파일명': filename
                            })
                except Exception as e:
                    print(f"게시글 {i}: 첨부파일 정보를 추출 중 오류 발생: {e}")
                    traceback.print_exc()
        
                # 개인정보(수) 추출 (예: JSON 데이터에서)
                개인정보_수 = ''
                try:
                    span_attach = driver.find_element(By.ID, 'attach').text
                    개인정보_수_matches = re.findall(r'"개인정보":\s*"(\d+)"', span_attach)
                    if 개인정보_수_matches:
                        개인정보_수 = 개인정보_수_matches[0]
                except:
                    # 혹은 다른 방식으로 개인정보 수를 추출할 수 있습니다.
                    개인정보_수 = ''
        
                # 진행 구분 설정: '제목'에 '추출완료일' 포함 시 "다운 완료"
                if '추출완료일' in 제목:
                    진행_구분 = '다운 완료'
                else:
                    진행_구분 = ''
        
                # 데이터 저장
                data = {
                    '등록일': 등록일_text,
                    '법인명': extract_corporate_name(작성자_full),
                    '제목': 제목,
                    '작성자': 작성자,
                    '링크': driver.current_url,
                    '파일형식': ', '.join([f['파일형식'] for f in 파일_리스트]),
                    '파일 용량': ', '.join([f['파일용량'] for f in 파일_리스트]),
                    '고유식별정보(수)': '',  # 공백으로 저장
                    '개인정보(수)': 개인정보_수,
                    '진행 구분': 진행_구분
                }
                data_list.append(data)
                print(f"게시글 {i}: 데이터 추출 완료")
            except Exception as e:
                print(f"게시글 {i}: 데이터 추출 중 오류 발생: {e}")
                traceback.print_exc()
        
            # 창 닫기 및 원래 창으로 전환
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
        
            # 잠시 대기
            time.sleep(2)
        
        # 데이터프레임 생성
        df = pd.DataFrame(data_list)
        
        ######################################엑셀화##############################################
        
        if not df.empty:
            try:
                    # 기존 엑셀 파일 불러오기
                wb = load_workbook(excel_file)
                if worksheet_name not in wb.sheetnames:
                    print(f"워크시트 '{worksheet_name}'이(가) 존재하지 않습니다.")
                    driver.quit()
                    sys.exit()
                ws = wb[worksheet_name]
            
                # 'S' 열(등록일)에서 데이터가 있는 마지막 행 찾기
                last_row = ws.max_row
                while last_row >= 5:  # 데이터가 시작되는 행 번호는 5
                    if ws.cell(row=last_row, column=19).value is not None:  # S열 (등록일) 확인
                        break
                    last_row -= 1
            
                # 새로운 데이터 입력 시작 행
                if last_row < 5:
                    start_row = 5  # 데이터 시작 행
                else:
                    start_row = last_row + 1
            
                # 데이터프레임의 열 순서 조정 (엑셀의 열 순서와 일치하도록)
                df = df[['등록일', '법인명', '제목', '작성자', '링크', '파일형식', '파일 용량', '고유식별정보(수)', '개인정보(수)', '진행 구분']]
            
                # 열 매핑 설정 (데이터프레임 열 이름과 엑셀 열 인덱스 매핑)
                column_mapping = {
                    '등록일': 19,          # S
                    '법인명': 20,          # T
                    '제목': 21,            # U
                    '작성자': 22,          # V
                    '링크': 23,            # W
                    '파일형식': 24,        # X
                    '파일 용량': 25,       # Y
                    '고유식별정보(수)': 26, # Z
                    '개인정보(수)': 27,    # AA
                    '진행 구분': 28        # AB
                }
            
                # 데이터프레임을 엑셀 워크시트에 쓰기
                for idx, row in df.iterrows():
                    # 각 열에 데이터 입력
                    for col_name, col_idx in column_mapping.items():
                        value = row[col_name]
                        ws.cell(row=start_row, column=col_idx, value=value)
                    start_row += 1
            
                # 엑셀 파일 저장
                wb.save(excel_file)
                print(f"데이터가 성공적으로 '{excel_file}' 파일에 저장되었습니다.")

            except Exception as e:
                print("엑셀 파일 처리 중 오류가 발생했습니다.")
                print(e)
                traceback.print_exc()
        else:
            print("추출된 데이터가 없습니다.")
    
    except Exception as e:
        print("스크립트 실행 중 예상치 못한 오류가 발생했습니다.")
        print(e)
        traceback.print_exc()
    finally:
        # 브라우저 종료
        driver.quit()
        print("브라우저가 종료되었습니다.")

if __name__ == "__main__":
    main()
