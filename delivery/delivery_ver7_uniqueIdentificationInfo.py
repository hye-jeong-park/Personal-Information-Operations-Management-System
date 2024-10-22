import re
import sys
import time
import traceback
import getpass
import logging
from typing import Tuple, Optional, List, Dict

import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook

# 설정
CRAWL_LIMIT = 21
EXCEL_FILE = r'C:\Users\PHJ\output\개인정보 운영대장.xlsx'
WORKSHEET_NAME = '개인정보 추출 및 이용 관리'

# 로깅 설정
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def extract_corporate_name(full_text: str) -> str:
    """
    법인명 추출: "게임사업3본부 K사업팀 / 홍길동님" 중 "게임사업3본부"만 추출
    """
    if '/' in full_text:
        return full_text.split('/')[0].strip().split()[0]
    return full_text.strip().split()[0]

def extract_file_info(file_info: str) -> Tuple[str, str]:
    """
    파일형식 및 파일 용량 추출
    """
    file_match = re.match(r'(.+?)\s*(?:&|[(])\s*([\d,\.]+\s*[KMGT]?B)', file_info, re.IGNORECASE)
    if file_match:
        filename_part = file_match.group(1).strip()
        size_part = file_match.group(2).strip()
    else:
        filename_part = file_info.strip()
        size_match = re.search(r'([\d,\.]+\s*[KMGT]?B)', filename_part, re.IGNORECASE)
        if size_match:
            size_part = size_match.group(1).strip()
            filename_part = filename_part.replace(size_part, '').strip()
        else:
            size_part = ''

    file_type = ''
    if '.zip' in filename_part.lower():
        file_type = 'Zip'
    elif '.xlsx' in filename_part.lower():
        file_type = 'Excel'

    size_match = re.match(r'([\d,\.]+)\s*([KMGT]?B)', size_part, re.IGNORECASE)
    if size_match:
        size_numeric = size_match.group(1).replace(',', '')
        size_unit = size_match.group(2).upper()
        file_size = f"{size_numeric} {size_unit}"
    else:
        file_size = size_part

    return file_type, file_size

def find_section_text(driver: webdriver.Chrome, section_titles: List[str]) -> Optional[str]:
    """
    특정 섹션의 제목을 기반으로 해당 섹션의 내용을 추출하는 함수
    """
    try:
        tr_elements = driver.find_elements(By.XPATH, '//table//tr')
        for tr in tr_elements:
            try:
                td_elements = tr.find_elements(By.TAG_NAME, 'td')
                if len(td_elements) >= 2:
                    header_text = ''.join([span.text.strip() for span in td_elements[0].find_elements(By.TAG_NAME, 'span')])

                    for section_title in section_titles:
                        if section_title in header_text:
                            return td_elements[1].text.strip()
            except Exception:
                continue
        return None
    except Exception as e:
        logging.error(f"find_section_text 오류: {e}")
        return None

def extract_attachment_info(driver: webdriver.Chrome) -> Tuple[str, str]:
    """
    메인 문서 내의 첨부파일 정보를 추출하는 함수
    """
    파일형식, 파일용량 = '', ''

    try:
        attm_read_div = driver.find_element(By.ID, 'attmRead')
        logging.info("첨부파일 div 찾음: attmRead")

        try:
            size_text = attm_read_div.find_element(By.XPATH, './/span[@class="attm-size"]').text.strip()
            size_match = re.match(r'([\d,\.]+)\s*([KMGT]?B)', size_text, re.IGNORECASE)
            if size_match:
                size_numeric = size_match.group(1).replace(',', '')
                size_unit = size_match.group(2).upper()
                파일용량 = f"{size_numeric} {size_unit}"
            else:
                파일용량 = size_text
            logging.info(f"파일용량 추출: {파일용량}")
        except Exception as e:
            logging.warning(f"파일용량 추출 중 오류 발생: {e}")

        try:
            filename = attm_read_div.find_element(By.XPATH, './/ul[contains(@class, "attm-list")]/li/a/strong').text.strip()
            if '.zip' in filename.lower():
                파일형식 = 'Zip'
            elif '.xlsx' in filename.lower():
                파일형식 = 'Excel'
            logging.info(f"파일형식 추출: {파일형식}")
        except Exception as e:
            logging.warning(f"파일형식 추출 중 오류 발생: {e}")
            파일형식 = ''
    except Exception as e:
        logging.warning(f"attmRead를 찾을 수 없음: {e}")

    if not 파일형식 and not 파일용량:
        try:
            iframe = driver.find_element(By.ID, 'ifa_form')
            driver.switch_to.frame(iframe)
            logging.info("iframe으로 전환하여 파일 정보 추출 시도")
            file_text = find_section_text(driver, ['파밀명 및 용량 (KB)', '파일명 및 용량 (KB)'])
            if file_text:
                logging.info(f"iframe 내에서 파일 정보 추출 시작: {file_text}")
                파일형식, 파일용량 = extract_file_info(file_text)
                logging.info(f"iframe 내에서 파일 정보 추출 완료: {파일형식}, {파일용량}")
            else:
                logging.warning("iframe 내에서 파일 정보 섹션을 찾을 수 없습니다.")
            driver.switch_to.default_content()
        except Exception as e:
            logging.error(f"iframe에서 파일 정보 추출 중 오류 발생: {e}")
            driver.switch_to.default_content()

    return 파일형식, 파일용량

def initialize_webdriver() -> webdriver.Chrome:
    """
    웹드라이버 초기화
    """
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    # options.add_argument("--headless")  # 필요 시 헤드리스 모드 활성화
    driver = webdriver.Chrome(options=options)
    return driver

def login(driver: webdriver.Chrome, username: str, password: str) -> bool:
    """
    로그인 처리
    """
    try:
        driver.get('https://gw.com2us.com/')
        username_input = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, 'Username'))
        )
        password_input = driver.find_element(By.ID, 'Password')

        username_input.send_keys(username)
        password_input.send_keys(password)

        login_button = driver.find_element(By.CLASS_NAME, 'btnLogin')
        login_button.click()

        WebDriverWait(driver, 20).until(
            EC.url_changes('https://gw.com2us.com/')
        )
        current_url = driver.current_url
        logging.info(f"로그인 후 현재 URL: {current_url}")

        if 'login' in current_url.lower():
            logging.error("로그인에 실패하였습니다.")
            return False
        return True
    except Exception as e:
        logging.error("로그인 성공 여부를 확인할 수 없습니다.")
        logging.error(e)
        traceback.print_exc()
        return False

def navigate_to_target_page(driver: webdriver.Chrome) -> bool:
    """
    개인정보 파일 전송 페이지로 이동
    """
    try:
        driver.get('https://gw.com2us.com/emate_app/00001/bbs/b2307140306.nsf/view?readform&viewname=view01')
        logging.info(f"페이지 이동 후 현재 URL: {driver.current_url}")
        return True
    except Exception as e:
        logging.error("타겟 페이지로 이동 중 오류 발생.")
        logging.error(e)
        traceback.print_exc()
        return False

def fetch_posts(driver: webdriver.Chrome) -> List[webdriver.remote.webelement.WebElement]:
    """
    게시글 목록 가져오기
    """
    try:
        posts = WebDriverWait(driver, 20).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'tr[class*="dhx_skyblue"]'))
        )
        total_posts = len(posts)
        logging.info(f"총 게시글 수: {total_posts}")
        return posts
    except Exception as e:
        logging.error("게시글 목록을 가져오는 중 오류 발생.")
        logging.error(e)
        traceback.print_exc()
        return []

def extract_post_data(driver: webdriver.Chrome, post: webdriver.remote.webelement.WebElement, index: int) -> Optional[Dict]:
    """
    게시글에서 데이터 추출
    """
    try:
        tds = post.find_elements(By.TAG_NAME, 'td')

        # 등록일
        if len(tds) >= 5:
            등록일_text = tds[4].get_attribute('title').strip() if tds[4].get_attribute('title') else tds[4].text.strip()
        else:
            logging.warning(f"게시글 {index}: 등록일 정보가 부족합니다.")
            등록일_text = ''

        # 작성자
        if len(tds) >= 3:
            작성자_td = tds[2]
            작성자 = 작성자_td.find_element(By.TAG_NAME, 'span').text.strip() if 작성자_td.find_elements(By.TAG_NAME, 'span') else 작성자_td.text.strip()
        else:
            logging.warning(f"게시글 {index}: 작성자 정보가 부족합니다.")
            작성자 = ''

        # 스크롤 및 클릭
        driver.execute_script("arguments[0].scrollIntoView();", post)
        WebDriverWait(driver, 20).until(EC.element_to_be_clickable(post))
        post.click()

        # 새 창으로 전환
        WebDriverWait(driver, 20).until(EC.number_of_windows_to_be(2))
        driver.switch_to.window(driver.window_handles[-1])
        logging.info(f"게시글 {index}: 새 창으로 전환")

        # 상세 페이지 로딩
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, 'HeaderTable'))
        )
        logging.info(f"게시글 {index}: 상세 페이지 로딩 완료")

        # 제목, 작성자, 등록일 상세
        제목 = driver.find_element(By.ID, 'DisSubject').text.strip() if driver.find_elements(By.ID, 'DisSubject') else ''
        작성자_full = driver.find_element(By.ID, 'DismyName').text.strip() if driver.find_elements(By.ID, 'DismyName') else ''
        등록일_text_detail = driver.find_element(By.ID, 'DiscDate').text.strip() if driver.find_elements(By.ID, 'DiscDate') else ''

        # 첨부파일 정보
        파일형식, 파일용량 = extract_attachment_info(driver)

        # iframe 전환
        법인명, 개인정보_수, 고유식별정보_수, 수신자 = '', 0, 0, ''
        try:
            iframe = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, 'ifa_form'))
            )
            driver.switch_to.frame(iframe)
            logging.info(f"게시글 {index}: iframe으로 전환")

            recipient_text = find_section_text(driver, ['수신자 (부서, 이름)', "Recipient's Department and Name"])
            if recipient_text:
                수신자 = recipient_text.strip()
                법인명 = extract_corporate_name(recipient_text)
                logging.info(f"게시글 {index}: 수신자 정보 추출 완료: {법인명}")
            else:
                logging.warning(f"게시글 {index}: 수신자 정보를 찾을 수 없습니다.")

            item_text = find_section_text(driver, ['추출된 항목 및 건수', 'Items and Counts Extracted'])
            if item_text:
                lines = item_text.strip().split('\n')
                keywords = ["주민등록번호", "여권번호", "운전면허의 면허번호", "외국인등록번호", "신분증"]
                found_keywords = False
                for line in lines:
                    line = line.strip()
                    # 전체 건수 추출
                    count_match = re.search(r'(\d{1,3}(?:,\d{3})*)\s*건', line)
                    if count_match:
                        count = int(count_match.group(1).replace(',', ''))
                        개인정보_수 += count
                    else:
                        count = 0  # 건수가 없는 경우 0으로 처리
                    # 키워드 포함 여부 확인
                    if any(keyword in line for keyword in keywords):
                        고유식별정보_수 += count
                        found_keywords = True
                if not found_keywords:
                    logging.info(f"게시글 {index}: 고유식별정보 미포함")
                else:
                    logging.info(f"게시글 {index}: 고유식별정보 수 추출 완료: {고유식별정보_수}")
                logging.info(f"게시글 {index}: 개인정보 수 추출 완료: {개인정보_수}")
            else:
                logging.warning(f"게시글 {index}: '추출된 항목 및 건수' 섹션을 찾을 수 없습니다.")

            driver.switch_to.default_content()
        except Exception as e:
            logging.error(f"게시글 {index}: iframe에서 데이터 추출 중 오류 발생: {e}")
            traceback.print_exc()
            driver.switch_to.default_content()

        # 진행 구분 설정
        진행_구분 = ''
        try:
            # 첨부파일 이력조회 버튼 클릭
            attm_log_button = driver.find_element(By.XPATH, '//a[span[text()="첨부파일 이력조회"]]')
            attm_log_button.click()
            logging.info(f"게시글 {index}: 첨부파일 이력조회 버튼 클릭")

            try:
                # 이력 테이블이 나타날 때까지 대기
                WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.XPATH, '//table[@id="ResultTable"]/tbody/tr'))
                )
                logging.info(f"게시글 {index}: 첨부파일 이력 테이블 로딩 완료")
            except Exception as e:
                logging.error(f"게시글 {index}: 첨부파일 이력 테이블 로딩 중 오류 발생: {e}")
                traceback.print_exc()
                return None

            # 이력 테이블에서 행들을 가져옴
            rows = driver.find_elements(By.XPATH, '//table[@id="ResultTable"]/tbody/tr')
            다운로드_이력_존재 = False

            for row in rows:
                cells = row.find_elements(By.TAG_NAME, 'td')
                if len(cells) >= 6:
                    구분 = cells[0].text.strip()
                    수행자_element = cells[1]
                    # 수행자 이름 추출
                    수행자 = 수행자_element.find_element(By.CLASS_NAME, 'pob').text.strip()
                    if 구분 == '다운로드' and 수행자 in 수신자:
                        다운로드_이력_존재 = True
                        logging.info(f"게시글 {index}: 다운로드 이력 발견 - 수행자: {수행자}")
                        break

            if 다운로드_이력_존재:
                진행_구분 = '다운 완료'
            else:
                진행_구분 = ''
        except Exception as e:
            logging.error(f"게시글 {index}: 첨부파일 이력조회 처리 중 오류 발생: {e}")
            traceback.print_exc()

        # 데이터 구성
        data = {
            '등록일': 등록일_text or 등록일_text_detail,
            '법인명': 법인명,
            '제목': 제목,
            '작성자': 작성자_full,
            '링크': driver.current_url,
            '파일형식': 파일형식,
            '파일 용량': 파일용량,
            '고유식별정보(수)': 고유식별정보_수,
            '개인정보(수)': 개인정보_수,
            '진행 구분': 진행_구분
        }

        logging.info(f"게시글 {index}: 데이터 추출 완료")
        return data

    except Exception as e:
        logging.error(f"게시글 {index}: 데이터 추출 중 오류 발생: {e}")
        traceback.print_exc()
        return None
    finally:
        # 창 닫기 및 원래 창으로 전환
        try:
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
        except Exception as e:
            logging.error(f"게시글 {index}: 창 닫기 및 전환 중 오류 발생: {e}")
            traceback.print_exc()
        time.sleep(2)

def save_to_excel(data_list: List[Dict]) -> None:
    """
    데이터프레임을 엑셀에 저장
    """
    if not data_list:
        logging.info("추출된 데이터가 없습니다.")
        return

    df = pd.DataFrame(data_list)

    try:
        wb = load_workbook(EXCEL_FILE)
        if WORKSHEET_NAME not in wb.sheetnames:
            logging.error(f"워크시트 '{WORKSHEET_NAME}'이(가) 존재하지 않습니다.")
            sys.exit()

        ws = wb[WORKSHEET_NAME]
        last_row = ws.max_row
        while last_row >= 5 and ws.cell(row=last_row, column=19).value is None:
            last_row -= 1

        start_row = max(last_row + 1, 5)

        df = df[['등록일', '법인명', '제목', '작성자', '링크', '파일형식', '파일 용량', '고유식별정보(수)', '개인정보(수)', '진행 구분']]

        column_mapping = {
            '등록일': 19,           # S
            '법인명': 20,           # T
            '제목': 21,             # U
            '작성자': 22,           # V
            '링크': 23,             # W
            '파일형식': 24,         # X
            '파일 용량': 25,        # Y
            '고유식별정보(수)': 26, # Z
            '개인정보(수)': 27,     # AA
            '진행 구분': 28         # AB
        }

        for _, row in df.iterrows():
            for col_name, col_idx in column_mapping.items():
                ws.cell(row=start_row, column=col_idx, value=row[col_name])
            start_row += 1

        wb.save(EXCEL_FILE)
        logging.info(f"데이터가 성공적으로 '{EXCEL_FILE}' 파일에 저장되었습니다.")

    except Exception as e:
        logging.error("엑셀 파일 처리 중 오류가 발생했습니다.")
        logging.error(e)
        traceback.print_exc()

def main() -> None:
    driver = initialize_webdriver()

    try:
        username = input('아이디를 입력하세요: ')
        password = getpass.getpass('비밀번호를 입력하세요: ')

        if not login(driver, username, password):
            driver.quit()
            sys.exit()

        if not navigate_to_target_page(driver):
            driver.quit()
            sys.exit()

        posts = fetch_posts(driver)
        total_posts = len(posts)

        if total_posts <= 1:
            logging.info("처리할 게시글이 없습니다. (첫 번째 게시글만 존재)")
            driver.quit()
            sys.exit()

        limit = min(CRAWL_LIMIT, total_posts - 1)
        logging.info(f"크롤링할 게시글 개수: {limit}")

        data_list = []

        for i in range(1, limit + 1):
            posts = driver.find_elements(By.CSS_SELECTOR, 'tr[class*="dhx_skyblue"]')
            if i >= len(posts):
                logging.warning(f"게시글 {i}은 존재하지 않습니다. 종료합니다.")
                break
            post = posts[i]

            data = extract_post_data(driver, post, i)
            if data:
                data_list.append(data)

        save_to_excel(data_list)

    except Exception as e:
        logging.error("스크립트 실행 중 예상치 못한 오류가 발생했습니다.")
        logging.error(e)
        traceback.print_exc()
    finally:
        driver.quit()
        logging.info("브라우저가 종료되었습니다.")

if __name__ == "__main__":
    main()